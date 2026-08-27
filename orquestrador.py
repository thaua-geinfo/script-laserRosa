#!/usr/bin/env python3
"""Orquestrador dos scripts de importacao do projeto Laser Rosa.

Coloque este arquivo junto aos scripts do projeto e execute:

    py orquestrador.py

Ordem padrao:
    1. Clientes
    2. Venda de Planos e Saldos
    3. Contas a Receber
    4. Contas a Pagar
    5. Agendamentos

O orquestrador:
- localiza uma raiz existente que contenha entrada e saida/saida com acento;
- nao cria as pastas principais do projeto;
- prioriza sempre os nomes canonicos dos scripts na raiz;
- aceita nomes canonicos e copias com sufixos como _ATUALIZADO, _CORRIGIDO e (2);
- executa cada etapa em um processo separado, usando o mesmo Python;
- captura a saida completa dos scripts para o log e mantem o terminal resumido;
- sobrescreve sempre o mesmo logOrquestrador.txt diretamente na raiz;
- no log, registra somente o resumo das etapas que tiveram sucesso e detalha a saida completa apenas em erros/bloqueios;
- qualquer arquivo/planilha de entrada ausente torna somente aquela etapa NAO_APLICAVEL;
- dependencias ausentes tornam a etapa dependente NAO_APLICAVEL sem executar dados incompletos;
- continua automaticamente as etapas independentes apos falhas/bloqueios;
- --parar-em-erro permite interromper deliberadamente na primeira pendencia real;
- mantem Agendamentos como ultima etapa da execucao completa.
"""

from __future__ import annotations

import argparse
import ast
import csv
import io
import os
import re
import shlex
import subprocess
import sys
import time
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from functools import lru_cache
from typing import Iterable, Sequence, TextIO

VERSION = "2026-08-26.10"
FIXED_PAYMENT_CODE = "Conta (uso financeiro)"


# ---------------------------------------------------------------------------
# Definicao das etapas
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class StageSpec:
    key: str
    label: str
    script_prefixes: tuple[str, ...]
    canonical_names: tuple[str, ...]
    dependencies: tuple[str, ...] = ()


STAGES: tuple[StageSpec, ...] = (
    StageSpec(
        key="clientes",
        label="Clientes",
        script_prefixes=("scriptcliente",),
        canonical_names=("scriptCliente.py",),
    ),
    StageSpec(
        key="venda-planos",
        label="Venda de Planos e Saldos",
        script_prefixes=("scriptvendaplanosaldo",),
        canonical_names=("scriptVendaPlanoSaldo.py", "scriptVendaPlanoSaldos.py"),
        dependencies=("clientes",),
    ),
    StageSpec(
        key="contas-receber",
        label="Contas a Receber",
        script_prefixes=("scriptcontasreceber",),
        canonical_names=("scriptContasReceber.py",),
        dependencies=("clientes",),
    ),
    StageSpec(
        key="contas-pagar",
        label="Contas a Pagar",
        script_prefixes=("scriptcontaspagar",),
        canonical_names=("scriptContasPagar.py",),
    ),
    StageSpec(
        key="agendamentos",
        label="Agendamentos",
        script_prefixes=("scriptagendamento",),
        canonical_names=("scriptAgendamento.py",),
        dependencies=("clientes",),
    ),
)

STAGE_BY_KEY = {stage.key: stage for stage in STAGES}
STAGE_ORDER = {stage.key: index for index, stage in enumerate(STAGES)}


ALIASES_RAW = {
    "cliente": "clientes",
    "clientes": "clientes",
    "cli": "clientes",
    "venda": "venda-planos",
    "vendas": "venda-planos",
    "plano": "venda-planos",
    "planos": "venda-planos",
    "venda-plano": "venda-planos",
    "venda-planos": "venda-planos",
    "venda-plano-saldo": "venda-planos",
    "venda-planos-saldos": "venda-planos",
    "vendas-saldos": "venda-planos",
    "receber": "contas-receber",
    "contas-receber": "contas-receber",
    "conta-receber": "contas-receber",
    "cr": "contas-receber",
    "pagar": "contas-pagar",
    "contas-pagar": "contas-pagar",
    "conta-pagar": "contas-pagar",
    "cp": "contas-pagar",
    "agendamento": "agendamentos",
    "agendamentos": "agendamentos",
    "agenda": "agendamentos",
    "todos": "todos",
    "todas": "todos",
}


# ---------------------------------------------------------------------------
# Estruturas auxiliares
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ProjectLayout:
    root: Path
    input_dir: Path
    output_dir: Path


@dataclass
class ScriptChoice:
    path: Path
    version: str
    version_key: tuple[int, ...]
    source: str
    candidate_count: int = 1

    def supports(self, option: str) -> bool:
        options, _ = script_cli_metadata(self.source)
        return option in options

    def choices_for(self, option: str) -> frozenset[str]:
        _, choices_items = script_cli_metadata(self.source)
        return dict(choices_items).get(option, frozenset())


@dataclass
class StageResult:
    stage: StageSpec
    status: str
    return_code: int | None = None
    duration_seconds: float = 0.0
    script: Path | None = None
    detail: str = ""
    version: str = ""
    command: str = ""
    output_lines: list[str] = field(default_factory=list)


@dataclass
class Reporter:
    log_handle: TextIO | None = None
    _closed: bool = field(default=False, init=False)

    def line(self, message: str = "") -> None:
        print(message, flush=True)
        if self.log_handle is not None:
            self.log_handle.write(message + "\n")
            self.log_handle.flush()

    def raw(self, message: str) -> None:
        # Saida de subprocesso ja pode conter quebra de linha.
        print(message, end="", flush=True)
        if self.log_handle is not None:
            self.log_handle.write(message)
            self.log_handle.flush()

    def close(self) -> None:
        if self._closed:
            return
        self._closed = True
        if self.log_handle is not None:
            self.log_handle.close()


# ---------------------------------------------------------------------------
# Normalizacao e descoberta da raiz
# ---------------------------------------------------------------------------


def ascii_fold(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value))
    return "".join(character for character in text if not unicodedata.combining(character))


def compact(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", ascii_fold(value).casefold())


def normalized_token(value: str) -> str:
    text = ascii_fold(value).casefold().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def existing_layout_dir(root: Path, logical_name: str) -> Path | None:
    if not root.is_dir():
        return None
    wanted = compact(logical_name)
    try:
        children = list(root.iterdir())
    except OSError:
        return None
    for child in children:
        if child.is_dir() and compact(child.name) == wanted:
            return child.resolve()
    return None


def project_root_score(root: Path, input_dir: Path, output_dir: Path) -> int:
    score = 0
    try:
        input_files = [path for path in input_dir.iterdir() if path.is_file()]
    except OSError:
        input_files = []
    try:
        output_files = [path for path in output_dir.iterdir() if path.is_file()]
    except OSError:
        output_files = []
    try:
        root_files = [path for path in root.iterdir() if path.is_file()]
    except OSError:
        root_files = []

    input_names = [compact(path.stem) for path in input_files]
    output_names = [compact(path.stem) for path in output_files]
    root_names = [compact(path.stem) for path in root_files]

    if any(re.fullmatch(r"depara\d*", name) for name in input_names):
        score += 50
    score += min(sum(name.startswith("modeloimportacao") for name in input_names), 12) * 5
    if any(name.startswith("planilhatratadacliente") for name in output_names):
        score += 25
    score += min(sum(name.startswith("script") for name in root_names), 10) * 2
    score += min(len(input_files), 20)
    return score


def find_project_layout(start: Path) -> ProjectLayout:
    start = start.resolve()
    candidates: list[tuple[int, int, Path, Path, Path]] = []
    checked: list[Path] = []

    for root in (start, *start.parents):
        root = root.resolve()
        checked.append(root)
        input_dir = existing_layout_dir(root, "entrada")
        output_dir = existing_layout_dir(root, "saida")
        if input_dir is None or output_dir is None:
            continue
        score = project_root_score(root, input_dir, output_dir)
        # Maior pontuacao vence. Em empate, prefere a raiz mais externa para
        # evitar estruturas de teste antigas com entrada/saida vazias.
        candidates.append((score, -len(root.parts), root, input_dir, output_dir))

    if not candidates:
        inspected = "\n  - ".join(str(path) for path in checked)
        raise FileNotFoundError(
            "Nao encontrei uma raiz existente contendo simultaneamente as pastas "
            "'entrada' e 'saida' (ou 'saída'). Pastas verificadas:\n  - "
            + inspected
        )

    _, _, root, input_dir, output_dir = max(candidates, key=lambda item: (item[0], item[1]))
    return ProjectLayout(root=root, input_dir=input_dir, output_dir=output_dir)


# ---------------------------------------------------------------------------
# Descoberta e selecao dos scripts
# ---------------------------------------------------------------------------


@lru_cache(maxsize=32)
def script_cli_metadata(
    source: str,
) -> tuple[frozenset[str], tuple[tuple[str, frozenset[str]], ...]]:
    """Extrai opcoes reais declaradas via argparse, ignorando docstrings/comentarios."""
    options: set[str] = set()
    choices_by_option: dict[str, frozenset[str]] = {}
    try:
        tree = ast.parse(source)
    except SyntaxError:
        # Fallback conservador para copias incompletas ou com sintaxe antiga.
        declared = re.findall(
            r"add_argument\s*\(\s*['\"](--[a-zA-Z0-9-]+)['\"]",
            source,
        )
        return frozenset(declared), ()

    for node in ast.walk(tree):
        if not isinstance(node, ast.Call):
            continue
        function = node.func
        if not isinstance(function, ast.Attribute) or function.attr != "add_argument":
            continue

        flags = [
            argument.value
            for argument in node.args
            if isinstance(argument, ast.Constant)
            and isinstance(argument.value, str)
            and argument.value.startswith("--")
        ]
        if not flags:
            continue
        options.update(flags)

        raw_choices: object | None = None
        for keyword in node.keywords:
            if keyword.arg != "choices":
                continue
            try:
                raw_choices = ast.literal_eval(keyword.value)
            except (ValueError, TypeError, SyntaxError):
                raw_choices = None
            break

        if isinstance(raw_choices, (list, tuple, set, frozenset)):
            values = frozenset(str(value) for value in raw_choices)
            for flag in flags:
                choices_by_option[flag] = values

    return frozenset(options), tuple(sorted(choices_by_option.items()))


VERSION_RE = re.compile(r"(?m)^\s*VERSION\s*=\s*['\"]([^'\"]+)['\"]")
BACKUP_MARKERS = ("backup", "antigo", "old", "original", "rascunho")


def read_script_source(path: Path) -> str:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def version_key(value: str) -> tuple[int, ...]:
    numbers = tuple(int(part) for part in re.findall(r"\d+", value))
    return numbers


def inspect_script(path: Path) -> tuple[str, tuple[int, ...], str]:
    source = read_script_source(path)
    match = VERSION_RE.search(source)
    if match:
        version = match.group(1).strip()
        key = version_key(version)
    else:
        filename_numbers = version_key(path.stem)
        version = ".".join(str(number) for number in filename_numbers) if filename_numbers else "nao informada"
        key = filename_numbers
    return version, key, source


def is_stage_script(path: Path, stage: StageSpec) -> bool:
    if not path.is_file() or path.suffix.casefold() != ".py":
        return False
    stem = compact(path.stem)
    return any(stem.startswith(prefix) for prefix in stage.script_prefixes)


def unique_paths(paths: Iterable[Path]) -> list[Path]:
    result: list[Path] = []
    seen: set[Path] = set()
    for path in paths:
        try:
            resolved = path.resolve()
        except OSError:
            resolved = path
        if resolved not in seen:
            seen.add(resolved)
            result.append(resolved)
    return result


def candidate_sort_key(
    path: Path,
    stage: StageSpec,
    script_dir: Path,
    version_tuple: tuple[int, ...],
) -> tuple[object, ...]:
    normalized_name = compact(path.name)
    canonical = compact(path.name) in {compact(name) for name in stage.canonical_names}
    has_backup_marker = any(marker in normalized_name for marker in BACKUP_MARKERS)
    same_directory = path.parent.resolve() == script_dir.resolve()
    try:
        modified = path.stat().st_mtime_ns
    except OSError:
        modified = 0
    return (
        not has_backup_marker,
        canonical,
        same_directory,
        bool(version_tuple),
        version_tuple,
        modified,
        path.name.casefold(),
    )


def discover_script(
    stage: StageSpec,
    search_dirs: Sequence[Path],
    script_dir: Path,
    override: Path | None = None,
) -> ScriptChoice | None:
    if override is not None:
        path = override.resolve()
        if not path.is_file():
            raise FileNotFoundError(f"Script informado para {stage.label} nao existe: {path}")
        if path.suffix.casefold() != ".py":
            raise ValueError(f"O arquivo informado para {stage.label} nao e um .py: {path}")
        version, key, source = inspect_script(path)
        return ScriptChoice(path=path, version=version, version_key=key, source=source, candidate_count=1)

    # Nomes canonicos oficiais sempre vencem copias com sufixo/versionadas.
    for directory in unique_paths(search_dirs):
        if not directory.is_dir():
            continue
        for canonical_name in stage.canonical_names:
            canonical_path = directory / canonical_name
            if canonical_path.is_file():
                version, key, source = inspect_script(canonical_path)
                return ScriptChoice(
                    path=canonical_path.resolve(),
                    version=version,
                    version_key=key,
                    source=source,
                    candidate_count=1,
                )

    candidates: list[Path] = []
    for directory in unique_paths(search_dirs):
        if not directory.is_dir():
            continue
        try:
            files = list(directory.iterdir())
        except OSError:
            continue
        candidates.extend(path for path in files if is_stage_script(path, stage))

    candidates = unique_paths(candidates)
    if not candidates:
        return None

    inspected: list[tuple[Path, str, tuple[int, ...], str]] = []
    for path in candidates:
        try:
            version, key, source = inspect_script(path)
        except OSError:
            continue
        inspected.append((path, version, key, source))

    if not inspected:
        return None

    path, version, key, source = max(
        inspected,
        key=lambda item: candidate_sort_key(item[0], stage, script_dir, item[2]),
    )
    return ScriptChoice(
        path=path,
        version=version,
        version_key=key,
        source=source,
        candidate_count=len(inspected),
    )


# ---------------------------------------------------------------------------
# Interpretacao da linha de comando
# ---------------------------------------------------------------------------


def build_aliases() -> dict[str, str]:
    aliases: dict[str, str] = {}
    for alias, target in ALIASES_RAW.items():
        aliases[normalized_token(alias)] = target
    for stage in STAGES:
        aliases[normalized_token(stage.key)] = stage.key
        aliases[normalized_token(stage.label)] = stage.key
    return aliases


ALIASES = build_aliases()


def resolve_stage_name(value: str) -> str:
    token = normalized_token(value)
    target = ALIASES.get(token)
    if target is None:
        valid = ", ".join(stage.key for stage in STAGES)
        raise ValueError(f"Etapa desconhecida: {value!r}. Etapas validas: {valid}.")
    return target


def expand_dependencies(keys: Iterable[str]) -> set[str]:
    expanded: set[str] = set()

    def add(key: str) -> None:
        if key in expanded:
            return
        for dependency in STAGE_BY_KEY[key].dependencies:
            add(dependency)
        expanded.add(key)

    for key in keys:
        add(key)
    return expanded


def select_stages(values: Sequence[str], without_dependencies: bool) -> list[StageSpec]:
    if not values:
        selected_keys = {stage.key for stage in STAGES}
    else:
        resolved = [resolve_stage_name(value) for value in values]
        if "todos" in resolved:
            selected_keys = {stage.key for stage in STAGES}
        else:
            selected_keys = set(resolved)

    if not without_dependencies:
        selected_keys = expand_dependencies(selected_keys)

    return [stage for stage in STAGES if stage.key in selected_keys]


def resolve_override_path(value: str, layout: ProjectLayout, script_dir: Path) -> Path:
    path = Path(value).expanduser()
    if path.is_absolute():
        return path.resolve()

    candidates = unique_paths((script_dir / path, layout.root / path))
    for candidate in candidates:
        if candidate.is_file():
            return candidate.resolve()
    # Mantem a localizacao mais intuitiva na mensagem de erro.
    return (layout.root / path).resolve()


def parse_overrides(
    values: Sequence[str],
    layout: ProjectLayout,
    script_dir: Path,
) -> dict[str, Path]:
    result: dict[str, Path] = {}
    for raw in values:
        if "=" not in raw:
            raise ValueError(
                f"Valor invalido em --script: {raw!r}. Use ETAPA=ARQUIVO, "
                "por exemplo: --script contas-pagar=scriptContasPagar.py"
            )
        stage_raw, path_raw = raw.split("=", 1)
        stage_key = resolve_stage_name(stage_raw)
        if stage_key == "todos":
            raise ValueError("--script exige uma etapa especifica; 'todos' nao e aceito.")
        if not path_raw.strip():
            raise ValueError(f"Caminho vazio em --script {raw!r}.")
        result[stage_key] = resolve_override_path(path_raw.strip(), layout, script_dir)
    return result


def normalize_import_date(value: str | None) -> str | None:
    if value is None:
        return None
    for date_format in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(value.strip(), date_format)
            return parsed.strftime("%d/%m/%Y")
        except ValueError:
            continue
    raise ValueError("--data-importacao deve usar DD/MM/AAAA ou AAAA-MM-DD.")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Executa os scripts de importacao do projeto Laser Rosa em ordem segura.",
        formatter_class=argparse.RawTextHelpFormatter,
        epilog=(
            "Exemplos:\n"
            "  py orquestrador.py\n"
            "  py orquestrador.py clientes contas-receber\n"
            "  py orquestrador.py agendamentos --sem-dependencias\n"
            "  py orquestrador.py --listar\n"
            "  py orquestrador.py --simular\n"
            "  py orquestrador.py --data-importacao 22/08/2026\n"
            "  py orquestrador.py --parar-em-erro\n"
            "  py orquestrador.py --script contas-pagar=scriptContasPagar.py"
        ),
    )
    parser.add_argument(
        "etapas",
        nargs="*",
        metavar="ETAPA",
        help=(
            "Etapas a executar. Sem informar, executa todas.\n"
            "Valores: clientes, venda-planos, contas-receber, contas-pagar, agendamentos."
        ),
    )
    parser.add_argument(
        "--listar",
        action="store_true",
        help="Mostra a raiz, as etapas e os scripts selecionados, sem executar.",
    )
    parser.add_argument(
        "--simular",
        action="store_true",
        help="Mostra os comandos que seriam executados, sem executar os scripts.",
    )
    parser.add_argument(
        "--continuar-em-erro",
        action="store_true",
        help="Compatibilidade: continuar etapas independentes ja e o comportamento padrao.",
    )
    parser.add_argument(
        "--parar-em-erro",
        action="store_true",
        help="Interrompe a execucao no primeiro bloqueio ou falha.",
    )
    parser.add_argument(
        "--ignorar-ausentes",
        action="store_true",
        help="Pula etapas cujo script nao foi encontrado, em vez de abortar no pre-check.",
    )
    parser.add_argument(
        "--sem-dependencias",
        action="store_true",
        help="Executa somente as etapas informadas, sem incluir Clientes automaticamente.",
    )
    parser.add_argument(
        "--data-importacao",
        help=(
            "Data comum em DD/MM/AAAA ou AAAA-MM-DD. O parametro e repassado somente "
            "aos scripts que declaram --data-importacao; os demais usam sua regra interna."
        ),
    )
    parser.add_argument(
        "--script",
        action="append",
        default=[],
        metavar="ETAPA=ARQUIVO",
        help="Fixa manualmente o arquivo de uma etapa. Pode ser repetido.",
    )
    log_group = parser.add_mutually_exclusive_group()
    log_group.add_argument(
        "--log",
        type=Path,
        help="Arquivo de log. Nome relativo e gravado diretamente na raiz do projeto.",
    )
    log_group.add_argument(
        "--sem-log",
        action="store_true",
        help="Nao grava o log consolidado da execucao.",
    )
    parser.add_argument(
        "--versao",
        action="version",
        version=f"%(prog)s {VERSION}",
    )
    return parser


# ---------------------------------------------------------------------------
# Montagem e execucao dos comandos
# ---------------------------------------------------------------------------


def automatic_script_arguments(
    stage: StageSpec,
    choice: ScriptChoice,
    import_date: str | None,
) -> tuple[list[str], list[str]]:
    arguments: list[str] = []
    notes: list[str] = []

    if import_date and stage.key != "clientes":
        if choice.supports("--data-importacao"):
            arguments.extend(("--data-importacao", import_date))
        else:
            notes.append(
                f"{stage.label}: o script nao declara --data-importacao; usara sua regra interna de data."
            )

    # Evita qualquer pergunta antiga de forma de pagamento em Contas a Pagar.
    if stage.key == "contas-pagar":
        if choice.supports("--nao-interativo"):
            arguments.append("--nao-interativo")
        if choice.supports("--forma-pagamento-padrao"):
            arguments.extend(("--forma-pagamento-padrao", FIXED_PAYMENT_CODE))

    # Nas versoes atuais, automatico preserva codigos de origem de ate 6 digitos
    # e usa 100000+ somente quando necessario.
    if (
        stage.key == "venda-planos"
        and choice.supports("--codigo-venda")
        and "automatico" in choice.choices_for("--codigo-venda")
    ):
        arguments.extend(("--codigo-venda", "automatico"))

    return arguments, notes


def command_for_stage(
    stage: StageSpec,
    choice: ScriptChoice,
    import_date: str | None,
) -> tuple[list[str], list[str]]:
    automatic_args, notes = automatic_script_arguments(stage, choice, import_date)
    command = [sys.executable, "-u", str(choice.path), *automatic_args]
    return command, notes


def format_command(command: Sequence[str]) -> str:
    if os.name == "nt":
        return subprocess.list2cmdline(list(command))
    return shlex.join(command)


def resolve_log_path(argument: Path | None, layout: ProjectLayout) -> Path:
    # Usa sempre um unico log na raiz; open_reporter abre em modo "w",
    # portanto cada execucao sobrescreve integralmente a anterior.
    if argument is None:
        return (layout.root / "logOrquestrador.txt").resolve()

    expanded = argument.expanduser()
    if expanded.is_absolute():
        return expanded.resolve()
    return (layout.root / expanded.name).resolve()


def cleanup_old_logs(layout: ProjectLayout) -> None:
    for directory in unique_paths((layout.root, layout.output_dir)):
        if not directory.is_dir():
            continue
        for old_log in directory.glob("logOrquestrador_*.txt"):
            try:
                old_log.unlink()
            except OSError:
                pass


def _read_csv_text(path: Path) -> tuple[str, str, str]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "cp1252", "latin1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError(f"Nao foi possivel ler {path.name}")
    first = next((line for line in text.splitlines() if line.strip()), "")
    delimiter = ";" if ";" in first else ","
    return text, delimiter, encoding


VALID_CLIENT_UFS = {
    "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS",
    "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC",
    "SP", "SE", "TO",
}
INVALID_CLIENT_VALUES = {
    "buscando", "invaliddate", "undefined", "indefinido", "invalido",
    "null", "none", "nan", "nat", "na", "naoinformado", "seminformacao",
}


def _client_punctuation_only(value: object) -> bool:
    text = str(value or "").strip()
    return bool(text) and not any(ch.isalnum() for ch in text)


def _client_placeholder(value: object) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    return compact(text.rstrip(" .,:;!?_-–—()[]{}")) in INVALID_CLIENT_VALUES


def _client_garbage(value: object) -> bool:
    # Usado apenas quando o campo precisa conter dado real (ex.: Nome).
    return _client_punctuation_only(value) or _client_placeholder(value)


def _client_name(value: object) -> str:
    text = " ".join(str(value or "").split()).strip()
    if not text or _client_garbage(text):
        return ""
    text = re.sub(r"^[=+@]+", "", text).strip()
    text = re.sub(r"^[\-–—.,;:_!?#$%&*/|]+\s*", "", text).strip()
    return text if any(ch.isalpha() for ch in text) else ""


def _client_gender(value: object) -> str:
    key = compact(value)
    if key == "masculino":
        return "Masculino"
    if key == "feminino":
        return "Feminino"
    return ""


def _client_uf(value: object) -> str:
    text = str(value or "").strip().upper()
    return text if text in VALID_CLIENT_UFS else ""


def normalize_client_csv(path: Path, *, complementary: bool = False) -> int:
    """Sanitiza qualquer planilha de Cliente gerada sem alterar dados validos."""
    if not path.is_file():
        return 0
    try:
        text, delimiter, encoding = _read_csv_text(path)
        rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
        if not rows:
            return 0
        headers = rows[0]
        keys = [compact(header) for header in headers]
        origin_i = next((i for i, key in enumerate(keys) if key == "codigoorigem"), None)
        status_i = next((i for i, key in enumerate(keys) if key == "status" or key.startswith("statusleads")), None)
        type_i = next((i for i, key in enumerate(keys) if key in {"tipoorigem", "tipodeorigem"}), None)
        required_indexes = {
            i for i, header in enumerate(headers)
            if ("obrigatorio" in compact(header) or "obrigatório" in str(header).casefold())
            and "obrigatoriose" not in compact(header)
            and "obrigatoriono" not in compact(header)
        }
        changes = 0
        for row in rows[1:]:
            if len(row) < len(headers):
                row.extend([""] * (len(headers) - len(row)))
            for i, key in enumerate(keys):
                old = row[i]
                if key.startswith("sexo"):
                    new = _client_gender(old)
                elif key == "ufsigla" or key.startswith("uf"):
                    new = _client_uf(old)
                elif _client_punctuation_only(old):
                    # Obrigatorio: preserva como veio da extracao.
                    # Opcional: pontuacao/simbolos isolados podem ser removidos.
                    new = old.strip() if i in required_indexes else ""
                elif key.startswith("nome") and "origem" not in key:
                    new = _client_name(old)
                elif _client_placeholder(old):
                    new = ""
                else:
                    new = old.strip()
                if new != old:
                    row[i] = new
                    changes += 1
            if complementary:
                if origin_i is not None and row[origin_i].strip():
                    row[origin_i] = ""
                    changes += 1
                if status_i is not None and not row[status_i].strip():
                    row[status_i] = "Leads"
                    changes += 1
                if type_i is not None and not row[type_i].strip():
                    row[type_i] = "Parcerias"
                    changes += 1
        temp = path.with_name(path.name + ".tmp")
        with temp.open("w", encoding=encoding, newline="") as handle:
            csv.writer(handle, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL).writerows(rows)
        temp.replace(path)
        return changes
    except OSError:
        return 0


def normalize_primary_clients(output_dir: Path) -> int:
    return normalize_client_csv(output_dir / "planilhaTratadaCliente.csv", complementary=False)


def normalize_client_complement(output_dir: Path) -> int:
    return normalize_client_csv(output_dir / "planilhaTratadaClienteVendaPlano.csv", complementary=True)


def open_reporter(log_path: Path | None) -> Reporter:
    if log_path is None:
        return Reporter()
    log_path.parent.mkdir(parents=True, exist_ok=True)
    handle = log_path.open("w", encoding="utf-8", newline="")
    return Reporter(log_handle=handle)


def child_environment(layout: ProjectLayout, import_date: str | None) -> dict[str, str]:
    environment = os.environ.copy()
    environment["PYTHONUTF8"] = "1"
    environment["PYTHONIOENCODING"] = "utf-8"
    environment["LASER_ROSA_PROJECT_ROOT"] = str(layout.root)
    environment["LASER_ROSA_INPUT_DIR"] = str(layout.input_dir)
    environment["LASER_ROSA_OUTPUT_DIR"] = str(layout.output_dir)
    if import_date:
        environment["LASER_ROSA_DATA_IMPORTACAO"] = import_date
    return environment


def run_command(
    command: Sequence[str],
    layout: ProjectLayout,
    reporter: Reporter,
    import_date: str | None,
) -> tuple[int, list[str]]:
    process: subprocess.Popen[str] | None = None
    output_lines: list[str] = []
    try:
        process = subprocess.Popen(
            list(command),
            cwd=str(layout.root),
            env=child_environment(layout, import_date),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            bufsize=1,
        )
        assert process.stdout is not None
        for line in process.stdout:
            output_lines.append(line.rstrip("\r\n"))
        return process.wait(), output_lines
    except KeyboardInterrupt:
        if process is not None and process.poll() is None:
            process.terminate()
            try:
                process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                process.kill()
                process.wait()
        raise


def missing_required_input_reason(stage: StageSpec, output_lines: Sequence[str]) -> str:
    """Identifica falta de arquivo/planilha necessario para executar uma etapa.

    Regra do projeto: os arquivos podem chegar em momentos diferentes. Quando
    uma etapa nao possui algum arquivo de entrada necessario, ela nao deve ser
    tratada como erro nem bloquear outras etapas. Erros de processamento (por
    exemplo PermissionError/WinError 5) continuam sendo falhas reais.
    """
    folded = [ascii_fold(line).strip().casefold() for line in output_lines]
    missing_markers = (
        "nao encontrada", "nao encontrado", "nao encontrei", "nao localizado",
        "nao localizada", "ausente", "not found", "does not exist",
    )

    for line in folded:
        if not any(marker in line for marker in missing_markers):
            continue
        # Falhas de acesso/escrita nao sao ausencia de insumo.
        if any(token in line for token in ("permissionerror", "acesso negado", "winerror 5")):
            continue

        # Fonte principal da etapa.
        if "extracao" in line:
            return "extracao necessaria nao encontrada nesta unidade"

        # Mensagens conhecidas por etapa.
        stage_tokens = {
            "clientes": ("clientes*.xlsx", "clientes*.xls", "clientes*.csv"),
            "venda-planos": ("arquivo de vendas", "arquivo de sessoes", "vendas.xls", "vendas.xlsx", "sessoes.xls", "sessoes.xlsx"),
            "contas-receber": ("contas a receber",),
            "contas-pagar": ("contas a pagar", "relacao de plano de contas"),
            "agendamentos": ("sessao*.xlsx", "sessao*.xls", "sessao nao encontrada", "nao encontrei sessao"),
        }.get(stage.key, ())
        if any(token in line for token in stage_tokens):
            if "relacao de plano de contas" in line:
                return "relacao de plano de contas nao encontrada; etapa desconsiderada nesta execucao"
            return "arquivo de origem necessario nao encontrado nesta unidade"

        # Arquivos auxiliares obrigatorios da propria etapa. A ausencia tambem
        # significa que a etapa ainda nao pode ser executada nesta rodada.
        if "modelo" in line:
            return "modelo necessario nao encontrado; etapa desconsiderada nesta execucao"
        if "de-para" in line or "de para" in line:
            return "DE-PARA necessario nao encontrado; etapa desconsiderada nesta execucao"
        if "plano de contas" in line:
            return "relacao de plano de contas nao encontrada; etapa desconsiderada nesta execucao"
        if "planilhatratadacliente" in line or "planilha tratada de cliente" in line:
            return "planilha de clientes necessaria nao encontrada; etapa desconsiderada nesta execucao"

        # Caminho explicitamente apontando para a pasta de entrada + arquivo.
        if "entrada" in line and any(ext in line for ext in (".xlsx", ".xls", ".csv")):
            return "arquivo necessario da pasta entrada nao encontrado; etapa desconsiderada nesta execucao"

    return ""


def blocking_configuration_reason(stage: StageSpec, output_lines: Sequence[str]) -> str:
    """Mantido apenas para bloqueios de validacao que nao sejam arquivo ausente."""
    return ""


def classify_stage_status(stage: StageSpec, return_code: int, output_lines: Sequence[str]) -> str:
    if return_code == 0:
        return "OK"
    if missing_required_input_reason(stage, output_lines):
        return "NAO_APLICAVEL"
    folded = [ascii_fold(line).strip().casefold() for line in output_lines]
    if any(line.startswith("bloqueado:") for line in folded):
        return "BLOQUEADO"
    return "FALHA"


def _compact_highlight(line: str, layout: ProjectLayout) -> str:
    text = line.strip()
    for prefix in (str(layout.root), str(layout.output_dir), str(layout.input_dir)):
        text = text.replace(prefix, "<raiz>")
    # Caminhos finais nao agregam ao resumo; preserva apenas o nome do arquivo.
    text = re.sub(r"\s+->\s+.*$", "", text)
    text = re.sub(r"\s+para\s+(?:[A-Za-z]:[\\/]|/|<raiz>[\\/]).*$", "", text, flags=re.IGNORECASE)
    return text


def success_highlights(output_lines: Sequence[str], layout: ProjectLayout, limit: int = 6) -> list[str]:
    highlights: list[str] = []
    for line in output_lines:
        raw = line.strip()
        folded = ascii_fold(raw).casefold()
        if not raw:
            continue
        include = (
            folded.startswith("ok:")
            or folded.startswith("ok [")
            or folded.startswith("atencao:")
            or folded.startswith("aviso:")
            or folded.startswith("confirmadas:")
            or folded.startswith("fornecedores complementares:")
            or folded.startswith("ordem de importacao:")
            or folded.startswith("fora desta importacao")
            or (folded.startswith("clientes:") and any(token in folded for token in ("duplic", "complement")))
            or (
                folded.startswith("validacao:")
                and any(token in folded for token in ("aviso", "erro", "pendencia", "rejeit"))
                and "nenhum erro ou aviso" not in folded
                and "nenhuma linha rejeitada" not in folded
            )
        )
        if include:
            value = _compact_highlight(raw, layout)
            if value not in highlights:
                highlights.append(value)
        if len(highlights) >= limit:
            break
    return highlights


def validation_problem_summary(output_lines: Sequence[str], layout: ProjectLayout) -> list[str]:
    """Resume arquivos de validacao citados por uma etapa com problema.

    O log nao deve despejar milhares de linhas. Quando existe um XLSX/CSV de
    validacao, agrega Nivel/Tipo e mostra apenas as categorias mais relevantes.
    """
    candidates: list[Path] = []
    for line in output_lines:
        raw = line.strip()
        folded = ascii_fold(raw).casefold()
        if not folded.startswith("validacao:"):
            continue
        value = raw.split(":", 1)[1].strip().strip('"')
        if not value:
            continue
        path = Path(value)
        if not path.is_absolute():
            path = layout.output_dir / path
        if path.is_file() and path not in candidates:
            candidates.append(path)

    summaries: list[str] = []
    for path in candidates:
        try:
            counts: dict[tuple[str, str], int] = {}
            if path.suffix.casefold() == ".xlsx":
                from openpyxl import load_workbook
                workbook = load_workbook(path, read_only=True, data_only=True)
                try:
                    sheet = workbook["Relatorio Validacao"] if "Relatorio Validacao" in workbook.sheetnames else workbook[workbook.sheetnames[-1]]
                    iterator = sheet.iter_rows(values_only=True)
                    headers = next(iterator, None)
                    if not headers:
                        continue
                    header_map = {compact(value): index for index, value in enumerate(headers) if value is not None}
                    level_i = header_map.get("nivel")
                    type_i = header_map.get("tipo")
                    if level_i is None or type_i is None:
                        continue
                    for row in iterator:
                        level = str(row[level_i] if level_i < len(row) and row[level_i] is not None else "").strip()
                        category = str(row[type_i] if type_i < len(row) and row[type_i] is not None else "").strip()
                        if not level or not category or level.upper() == "INFO":
                            continue
                        key = (level.upper(), category)
                        counts[key] = counts.get(key, 0) + 1
                finally:
                    workbook.close()
            elif path.suffix.casefold() in {".csv", ".txt"}:
                text, delimiter, _ = _read_csv_text(path)
                rows = csv.reader(io.StringIO(text), delimiter=delimiter)
                headers = next(rows, None)
                if not headers:
                    continue
                header_map = {compact(value): index for index, value in enumerate(headers)}
                level_i = header_map.get("nivel")
                type_i = header_map.get("tipo") or header_map.get("motivodarejeicao")
                if type_i is None:
                    continue
                for row in rows:
                    level = str(row[level_i]).strip() if level_i is not None and level_i < len(row) else "ERRO"
                    category = str(row[type_i]).strip() if type_i < len(row) else ""
                    if not category or level.upper() == "INFO":
                        continue
                    key = (level.upper() or "ERRO", category)
                    counts[key] = counts.get(key, 0) + 1
            if counts:
                summaries.append(f"Arquivo de validacao: {path.name}")
                total = sum(counts.values())
                summaries.append(f"Ocorrencias relevantes: {total}")
                for (level, category), count in sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:15]:
                    summaries.append(f"  {level}: {category} = {count}")
                if len(counts) > 15:
                    summaries.append(f"  ... mais {len(counts) - 15} categoria(s)")
        except Exception as exc:
            summaries.append(f"Nao foi possivel resumir {path.name}: {type(exc).__name__}: {exc}")
    return summaries


def write_execution_log(
    log_path: Path | None,
    *,
    layout: ProjectLayout,
    results: Sequence[StageResult],
    total_duration: float,
    finished: bool,
) -> None:
    if log_path is None:
        return
    problem_statuses = {"FALHA", "BLOQUEADO", "INTERROMPIDO"}
    problems = [result for result in results if result.status in problem_statuses]
    if not finished:
        overall = "EM ANDAMENTO"
    elif problems:
        overall = "COM PENDENCIAS"
    elif any(result.status in {"PULADO", "NAO_APLICAVEL"} for result in results):
        overall = "OK / PARCIAL"
    else:
        overall = "OK"

    lines: list[str] = [
        f"Laser Rosa - Orquestrador v{VERSION}",
        f"Execucao: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        f"Raiz: {layout.root}",
        f"Resultado geral: {overall}",
        f"Duracao: {total_duration:.1f}s",
        "",
        "RESUMO",
    ]
    for result in results:
        status = result.status if result.status != "FALHA" else f"FALHA({result.return_code})"
        lines.append(f"{status:<12} | {result.stage.label:<30} | {result.duration_seconds:>7.1f}s | {result.script.name if result.script else '-'}")
        if result.status == "OK":
            for highlight in success_highlights(result.output_lines, layout):
                lines.append(f"  - {highlight}")
        elif result.status == "NAO_APLICAVEL":
            lines.append(f"  - {result.detail or 'arquivo necessario nao disponivel nesta execucao'}")
        elif result.detail:
            lines.append(f"  - {result.detail}")

    if problems:
        lines.extend(["", "DETALHES DOS PROBLEMAS"])
        for result in problems:
            status = result.status if result.status != "FALHA" else f"FALHA({result.return_code})"
            lines.extend([
                "",
                f"[{result.stage.label}] {status}",
                f"Script: {result.script if result.script else '-'}",
                f"Versao: {result.version or '-'}",
                f"Codigo de retorno: {result.return_code if result.return_code is not None else '-'}",
                f"Comando: {result.command or '-'}",
                "Saida completa da etapa:",
            ])
            validation_summary = validation_problem_summary(result.output_lines, layout)
            if validation_summary:
                lines.append("Resumo da validacao:")
                lines.extend(f"  {line}" for line in validation_summary)
                lines.append("Saida da etapa:")
            if result.output_lines:
                lines.extend(f"  {line}" for line in result.output_lines)
            else:
                lines.append("  (nenhuma saida capturada)")

    lines.append("")
    try:
        log_path.parent.mkdir(parents=True, exist_ok=True)
        log_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    except OSError as exc:
        print(f"AVISO: nao foi possivel atualizar o log {log_path}: {exc}", file=sys.stderr)


# ---------------------------------------------------------------------------
# Apresentacao e fluxo principal
# ---------------------------------------------------------------------------


def print_inventory(
    reporter: Reporter,
    layout: ProjectLayout,
    selected: Sequence[StageSpec],
    choices: dict[str, ScriptChoice | None],
) -> None:
    reporter.line(f"Laser Rosa - Orquestrador v{VERSION}")
    reporter.line(f"Raiz: {layout.root}")
    reporter.line(f"Entrada: {layout.input_dir}")
    reporter.line(f"Saida: {layout.output_dir}")
    reporter.line("Etapas selecionadas:")
    for index, stage in enumerate(selected, start=1):
        choice = choices.get(stage.key)
        dependency_text = ", ".join(stage.dependencies) if stage.dependencies else "nenhuma"
        if choice is None:
            reporter.line(
                f"  {index}. {stage.label}: NAO ENCONTRADO | dependencias: {dependency_text}"
            )
            continue
        multiple = (
            f" | {choice.candidate_count} candidatos; maior VERSION selecionado"
            if choice.candidate_count > 1
            else ""
        )
        reporter.line(
            f"  {index}. {stage.label}: {choice.path.name} | versao {choice.version}"
            f" | dependencias: {dependency_text}{multiple}"
        )


def print_summary(reporter: Reporter, results: Sequence[StageResult]) -> None:
    reporter.line("")
    reporter.line("RESUMO")
    for result in results:
        status = f"FALHA({result.return_code})" if result.status == "FALHA" else result.status
        duration = f"{result.duration_seconds:.1f}s" if result.duration_seconds else "-"
        detail = f" | {result.detail}" if result.status != "OK" and result.detail else ""
        reporter.line(f"{status:<13} | {result.stage.label:<28} | {duration:>7}{detail}")


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    script_dir = Path(__file__).resolve().parent

    try:
        layout = find_project_layout(script_dir)
        selected = select_stages(args.etapas, args.sem_dependencias)
        import_date = normalize_import_date(args.data_importacao)
        overrides = parse_overrides(args.script, layout, script_dir)
    except (FileNotFoundError, ValueError, OSError) as exc:
        print(f"ERRO: {exc}", file=sys.stderr)
        return 1

    search_dirs = unique_paths((script_dir, layout.root))
    choices: dict[str, ScriptChoice | None] = {}
    try:
        for stage in selected:
            choices[stage.key] = discover_script(
                stage,
                search_dirs=search_dirs,
                script_dir=script_dir,
                override=overrides.get(stage.key),
            )
    except (FileNotFoundError, ValueError, OSError) as exc:
        print(f"ERRO: {exc}", file=sys.stderr)
        return 1

    # --listar nao cria log e nao altera nenhum arquivo.
    if args.listar:
        reporter = Reporter()
        print_inventory(reporter, layout, selected, choices)
        return 0 if all(choices[stage.key] is not None for stage in selected) else 1

    missing = [stage for stage in selected if choices.get(stage.key) is None]
    if missing and not args.ignorar_ausentes:
        reporter = Reporter()
        print_inventory(reporter, layout, selected, choices)
        names = ", ".join(stage.label for stage in missing)
        reporter.line("")
        reporter.line(
            f"ERRO: script nao encontrado para: {names}. "
            "Use --ignorar-ausentes para pular ou --script ETAPA=ARQUIVO para fixar o caminho."
        )
        return 1

    log_path: Path | None = None
    if not args.sem_log and not args.simular:
        try:
            log_path = resolve_log_path(args.log, layout)
        except OSError as exc:
            print(f"ERRO ao definir log: {exc}", file=sys.stderr)
            return 1

    cleanup_old_logs(layout)

    # O terminal continua em tempo real. O log e escrito separadamente de forma
    # curada: resumo nos sucessos e saida completa apenas nos problemas.
    reporter = Reporter()

    results: list[StageResult] = []
    status_by_stage: dict[str, str] = {}
    notes_printed: set[str] = set()

    try:
        reporter.line(f"Laser Rosa - Orquestrador v{VERSION}")
        reporter.line(f"Raiz: {layout.root}")
        if import_date:
            reporter.line(f"Data da importacao: {import_date}")
        if log_path is not None:
            reporter.line(f"Log: {log_path}")

        if args.simular:
            reporter.line("")
            reporter.line("SIMULACAO - nenhum script sera executado")
            for stage in selected:
                choice = choices.get(stage.key)
                if choice is None:
                    reporter.line(f"  {stage.label}: PULADO - script ausente")
                    continue
                command, notes = command_for_stage(stage, choice, import_date)
                reporter.line(f"  {stage.label}: {format_command(command)}")
                for note in notes:
                    reporter.line(f"    AVISO: {note}")
            return 0 if not missing else 1

        overall_start = time.monotonic()
        reporter.line("")
        reporter.line("EXECUCAO")

        for position, stage in enumerate(selected, start=1):
            choice = choices.get(stage.key)

            if choice is None:
                result = StageResult(
                    stage=stage,
                    status="PULADO",
                    detail="script nao encontrado",
                )
                results.append(result)
                status_by_stage[stage.key] = result.status
                reporter.line("")
                reporter.line(f"PULADO [{position}/{len(selected)}] {stage.label}: script nao encontrado.")
                continue

            failed_dependencies = [
                dependency
                for dependency in stage.dependencies
                if dependency in status_by_stage and status_by_stage[dependency] != "OK"
            ]
            if failed_dependencies:
                dependency_labels = ", ".join(STAGE_BY_KEY[key].label for key in failed_dependencies)
                dependency_statuses = [status_by_stage[key] for key in failed_dependencies]
                dependency_unavailable = all(status == "NAO_APLICAVEL" for status in dependency_statuses)
                if dependency_unavailable:
                    dep_status = "NAO_APLICAVEL"
                    detail = f"dependencia nao disponivel nesta execucao: {dependency_labels}"
                    terminal_prefix = "NAO APLICAVEL"
                else:
                    dep_status = "PULADO"
                    detail = f"dependencia com pendencia: {dependency_labels}"
                    terminal_prefix = "PULADO"
                result = StageResult(
                    stage=stage,
                    status=dep_status,
                    script=choice.path,
                    detail=detail,
                )
                results.append(result)
                status_by_stage[stage.key] = result.status
                reporter.line("")
                reporter.line(
                    f"{terminal_prefix} [{position}/{len(selected)}] {stage.label}: {detail}."
                )
                continue

            command, notes = command_for_stage(stage, choice, import_date)
            for note in notes:
                if note not in notes_printed:
                    reporter.line(f"AVISO: {note}")
                    notes_printed.add(note)

            reporter.line(f"[{position}/{len(selected)}] {stage.label}...")

            started = time.monotonic()
            try:
                return_code, output_lines = run_command(command, layout, reporter, import_date)
            except KeyboardInterrupt:
                duration = time.monotonic() - started
                result = StageResult(
                    stage=stage,
                    status="INTERROMPIDO",
                    return_code=130,
                    duration_seconds=duration,
                    script=choice.path,
                    detail="execucao interrompida pelo usuario",
                    version=choice.version,
                    command=format_command(command),
                    output_lines=[],
                )
                results.append(result)
                status_by_stage[stage.key] = result.status
                reporter.line("")
                reporter.line("Execucao interrompida pelo usuario.")
                print_summary(reporter, results)
                write_execution_log(
                    log_path, layout=layout, results=results,
                    total_duration=time.monotonic() - overall_start, finished=True,
                )
                return 130

            duration = time.monotonic() - started
            classified_status = classify_stage_status(stage, return_code, output_lines)
            if classified_status == "OK":
                result = StageResult(
                    stage=stage,
                    status="OK",
                    return_code=0,
                    duration_seconds=duration,
                    script=choice.path,
                    version=choice.version,
                    command=format_command(command),
                    output_lines=output_lines,
                )
                highlights = success_highlights(output_lines, layout, limit=4)
                extra = " | " + " | ".join(highlights) if highlights else ""
                reporter.line(f"  OK | {duration:.1f}s{extra}")
                if stage.key == "clientes":
                    normalized = normalize_primary_clients(layout.output_dir)
                    if normalized:
                        reporter.line(f"Clientes: {normalized} celula(s) normalizada(s) pela validacao final.")
                if stage.key in {"venda-planos", "contas-receber", "agendamentos"}:
                    normalized = normalize_client_complement(layout.output_dir)
                    if normalized:
                        reporter.line(f"Clientes complementares: {normalized} ajuste(s) de normalizacao aplicado(s).")
            elif classified_status == "NAO_APLICAVEL":
                reason = missing_required_input_reason(stage, output_lines) or "arquivo necessario nao disponivel nesta execucao"
                result = StageResult(
                    stage=stage,
                    status="NAO_APLICAVEL",
                    return_code=return_code,
                    duration_seconds=duration,
                    script=choice.path,
                    detail=reason,
                    version=choice.version,
                    command=format_command(command),
                    output_lines=output_lines,
                )
                reporter.line(f"  NAO_APLICAVEL | {duration:.1f}s | {reason}")
            else:
                config_reason = blocking_configuration_reason(stage, output_lines)
                detail = (
                    "consulte a secao DETALHES DOS PROBLEMAS no log"
                    if classified_status == "FALHA"
                    else (config_reason or "etapa bloqueada por validacao; detalhes completos no log")
                )
                result = StageResult(
                    stage=stage,
                    status=classified_status,
                    return_code=return_code,
                    duration_seconds=duration,
                    script=choice.path,
                    detail=detail,
                    version=choice.version,
                    command=format_command(command),
                    output_lines=output_lines,
                )
                reporter.line(f"  {classified_status} | {duration:.1f}s | detalhes no log")

            results.append(result)
            status_by_stage[stage.key] = result.status
            write_execution_log(
                log_path, layout=layout, results=results,
                total_duration=time.monotonic() - overall_start, finished=False,
            )

            if classified_status in {"FALHA", "BLOQUEADO"}:
                if args.parar_em_erro:
                    reporter.line("Execucao interrompida por --parar-em-erro.")
                    break

        total_duration = time.monotonic() - overall_start
        print_summary(reporter, results)
        reporter.line(f"Duracao total: {total_duration:.1f}s")
        if log_path is not None:
            reporter.line(f"Log consolidado: {log_path}")
        write_execution_log(
            log_path, layout=layout, results=results,
            total_duration=total_duration, finished=True,
        )

        failed = any(result.status in {"FALHA", "BLOQUEADO", "INTERROMPIDO"} for result in results)
        missing_or_skipped = any(
            result.status == "PULADO" and "script nao encontrado" in result.detail
            for result in results
        )
        return 1 if failed or missing_or_skipped else 0
    finally:
        reporter.close()


if __name__ == "__main__":
    raise SystemExit(main())
