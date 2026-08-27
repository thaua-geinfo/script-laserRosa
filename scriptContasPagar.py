#!/usr/bin/env python3
"""Tratamento de Contas a Pagar - Laser Rosa.

Estrutura esperada:
    raiz/
      scriptContasPagar.py
      entrada/
        Contas a Pagar.xlsx
        Contas a Pagar - Plano de contas*.xlsx
        modeloImportacaoContasPagar.csv
        modeloImportacaoFornecedor.csv
      saida/
        planilhaTratadaFornecedor.csv        # fonte oficial, quando existente
        planilhaTratadaFornecedorContasPagar.csv
        planilhaTratadaContasPagar.csv
        validacaoContasPagar.xlsx            # somente quando houver pendencias

Regras aplicadas:
- modelos sao usados somente como cabecalho/formato; suas linhas sao ignoradas;
- a extracao Contas a Pagar e a unica fonte dos lancamentos;
- o arquivo Plano de contas e usado somente como relacao Id -> codigo/nome da conta;
- fornecedores sao relacionados pelas planilhas tratadas existentes; os ausentes
  sao criados em arquivo complementar a partir do modelo de fornecedor;
- a Observacao do fornecedor complementar usa o texto
  "Fornecedor gerado por Contas a Pagar | Importacao DD/MM/AAAA";
- em cada execucao, a data das observacoes dos fornecedores complementares
  e atualizada para a data efetivamente usada pela execucao;
- o campo Codigo Forma de Pagamento recebe sempre o valor fixo
  "Conta (uso financeiro)", conforme regra definida para esta importacao;
- o DE-PARA nao e consultado para a forma de pagamento de Contas a Pagar;
- o codigo do plano de contas conserva somente numeros e pontuacao ate o ultimo
  numero existente;
- nao cria automaticamente as pastas entrada/saida;
- remove aspas simples/duplas, barras invertidas, controles e quebras de linha;
- Observacao termina sempre com Importacao DD/MM/AAAA;
- arquivo de validacao existe apenas quando houver erro ou aviso;
- falhas por arquivo aberto nao geram validacao de dados;
- arquivos finais sao gravados de forma transacional; se estiverem abertos
  no Excel, o script preserva a versao anterior e informa o arquivo bloqueado;
- quando o conteudo gerado ja e identico ao CSV existente, a sobrescrita e
  dispensada, inclusive se o arquivo estiver aberto apenas para consulta.

Uso normal:
    py scriptContasPagar.py

Exemplo com parametros:
    py scriptContasPagar.py ^
      --entrada "Contas a Pagar.xlsx" ^
      --plano-contas "Contas a Pagar - Plano de contas.xlsx"
"""

from __future__ import annotations

import argparse
import errno
import hashlib
import csv
import os
import shutil
import time
import io
import re
import sys
import unicodedata
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable, Sequence
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

VERSION = "2026-08-25.1"
FIXED_PAYMENT_CODE = "Conta (uso financeiro)"
SUPPLIER_OBSERVATION_PREFIX = "Fornecedor gerado por Contas a Pagar"
LEGACY_SUPPLIER_OBSERVATION_PREFIXES = (
    "Fornecedor complementar gerado por Contas a Pagar",
)
NULL_WORDS = {"", "null", "none", "nan", "nat", "n/a", "na"}
FORBIDDEN_TRANSLATION = str.maketrans({'"': "", "'": " ", "\\": ""})

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR
INPUT_DIR = PROJECT_ROOT / "entrada"
OUTPUT_DIR = PROJECT_ROOT / "saida"


def script_hash() -> str:
    """Identificador curto para confirmar exatamente qual copia foi executada."""
    try:
        return hashlib.sha256(Path(__file__).read_bytes()).hexdigest()[:12]
    except OSError:
        return "indisponivel"


# ---------------------------------------------------------------------------
# Limpeza, normalizacao e formatacao
# ---------------------------------------------------------------------------

def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "S" if value else "N"
    text = unicodedata.normalize("NFKC", str(value))
    if text.strip().casefold() in NULL_WORDS:
        return ""
    text = text.translate(FORBIDDEN_TRANSLATION)
    text = "".join(
        character
        for character in text
        if unicodedata.category(character) not in {"Cc", "Cf"}
    )
    text = text.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    return " ".join(text.split()).strip()


def ascii_fold(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean_text(value))
    return "".join(character for character in text if not unicodedata.combining(character))


def norm(value: Any) -> str:
    text = ascii_fold(value).casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def compact(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", ascii_fold(value).casefold())


def exact_name_key(value: Any) -> str:
    return clean_text(value).casefold()


def normalized_name_key(value: Any) -> str:
    return norm(value)


def singular_name_key(value: Any) -> str:
    tokens: list[str] = []
    for token in norm(value).split():
        if len(token) > 3 and token.endswith("s"):
            token = token[:-1]
        tokens.append(token)
    return " ".join(tokens)


def digits_only(value: Any) -> str:
    return "".join(character for character in clean_text(value) if character.isdigit())


def normalize_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else format(value, ".15g")
    text = clean_text(value)
    if re.fullmatch(r"[+-]?\d+[.,]0+", text):
        return re.split(r"[.,]", text, 1)[0]
    return text


def parse_date(value: Any, *, date1904: bool = False) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        number = float(value)
        if 0 < number < 100000:
            base = date(1904, 1, 1) if date1904 else date(1899, 12, 30)
            return base + timedelta(days=int(number))
    text = clean_text(value)
    if not text:
        return None
    candidates = [
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d-%m-%Y",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
    ]
    for fmt in candidates:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})[T ]", text)
    if match:
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    return None


def date_out(value: Any, *, date1904: bool = False) -> str:
    parsed = parse_date(value, date1904=date1904)
    return parsed.strftime("%d/%m/%Y") if parsed else ""


def decimal_value(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None

    text = clean_text(value)
    if not text:
        return None
    negative = text.startswith("(") and text.endswith(")")
    text = text.replace("R$", "").replace("r$", "").replace(" ", "")
    text = text.strip("()")
    text = re.sub(r"[^0-9,\.\-+]", "", text)
    if not text:
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    elif "." in text:
        if re.fullmatch(r"[+-]?\d{1,3}(?:\.\d{3})+", text):
            text = text.replace(".", "")

    try:
        result = Decimal(text)
    except InvalidOperation:
        return None
    return -result if negative and result > 0 else result


def money(value: Any) -> str:
    parsed = decimal_value(value)
    if parsed is None:
        return ""
    quantized = parsed.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    return format(quantized, ".2f").replace(".", ",")


def plan_account_code(value: Any) -> str:
    """Mantem numeros e pontuacao e corta tudo depois do ultimo numero."""
    text = clean_text(value)
    filtered = "".join(
        character
        for character in text
        if character.isdigit() or character in ".,-/:"
    )
    last_digit = max((index for index, character in enumerate(filtered) if character.isdigit()), default=-1)
    return filtered[: last_digit + 1] if last_digit >= 0 else ""


def output_text(value: Any, delimiter: str = ";") -> str:
    text = clean_text(value)
    if delimiter:
        text = text.replace(delimiter, ",")
    return text


def encoding_safe(value: Any, encoding: str, delimiter: str = ";") -> str:
    text = output_text(value, delimiter)
    return text.encode(encoding, errors="replace").decode(encoding)


# ---------------------------------------------------------------------------
# Estruturas de dados
# ---------------------------------------------------------------------------

@dataclass
class Table:
    path: Path
    sheet: str
    header_row: int
    headers: list[Any]
    rows: list[tuple[int, list[Any]]]
    date1904: bool = False
    delimiter: str = ";"
    encoding: str = "utf-8-sig"


@dataclass
class RawSheet:
    path: Path
    name: str
    rows: list[tuple[int, list[Any]]]
    date1904: bool = False


@dataclass
class Issue:
    level: str
    category: str
    file: str = ""
    line: str = ""
    source_id: str = ""
    supplier: str = ""
    field: str = ""
    origin: str = ""
    compared: str = ""
    detail: str = ""


@dataclass
class ValidationReport:
    issues: list[Issue] = field(default_factory=list)

    def add(
        self,
        level: str,
        category: str,
        *,
        file: Path | str = "",
        line: int | str = "",
        source_id: Any = "",
        supplier: Any = "",
        field: Any = "",
        origin: Any = "",
        compared: Any = "",
        detail: Any = "",
    ) -> None:
        self.issues.append(
            Issue(
                level=level,
                category=clean_text(category),
                file=Path(file).name if file else "",
                line=str(line) if line != "" else "",
                source_id=normalize_id(source_id),
                supplier=clean_text(supplier),
                field=clean_text(field),
                origin=clean_text(origin),
                compared=clean_text(compared),
                detail=clean_text(detail),
            )
        )

    def error(self, category: str, **kwargs: Any) -> None:
        self.add("ERRO", category, **kwargs)

    def warning(self, category: str, **kwargs: Any) -> None:
        self.add("AVISO", category, **kwargs)

    @property
    def error_count(self) -> int:
        return sum(issue.level == "ERRO" for issue in self.issues)

    @property
    def warning_count(self) -> int:
        return sum(issue.level == "AVISO" for issue in self.issues)

    @property
    def findings_count(self) -> int:
        return self.error_count + self.warning_count


@dataclass(frozen=True)
class AccountRelation:
    code: str
    name: str
    source_line: int


@dataclass(frozen=True)
class PaymentRule:
    source_name: str
    code: str
    account_origin: str = ""


@dataclass
class SupplierCatalog:
    by_exact: dict[str, set[str]] = field(default_factory=lambda: defaultdict(set))
    by_normalized: dict[str, set[str]] = field(default_factory=lambda: defaultdict(set))
    by_singular: dict[str, set[str]] = field(default_factory=lambda: defaultdict(set))
    by_compact: dict[str, set[str]] = field(default_factory=lambda: defaultdict(set))
    used_codes: set[int] = field(default_factory=set)


@dataclass(frozen=True)
class GeneratedSupplier:
    code: str
    name: str


@dataclass
class ProcessResult:
    output_rows: list[list[str]]
    generated_suppliers: list[GeneratedSupplier]
    source_count: int
    confirmed_count: int
    unconfirmed_count: int
    source_total: Decimal
    output_total: Decimal


# ---------------------------------------------------------------------------
# Estrutura do projeto e descoberta de arquivos
# ---------------------------------------------------------------------------

def _existing_layout_dir(root: Path, logical_name: str) -> Path | None:
    if not root.is_dir():
        return None
    wanted = compact(logical_name)
    try:
        children = list(root.iterdir())
    except OSError:
        return None
    for child in children:
        if child.is_dir() and compact(child.name) == wanted:
            return child.resolve()
    return None


def configure_project_layout() -> tuple[Path, Path, Path]:
    global PROJECT_ROOT, INPUT_DIR, OUTPUT_DIR
    candidates: list[tuple[int, int, Path, Path, Path]] = []
    for distance, root in enumerate((SCRIPT_DIR, *SCRIPT_DIR.parents)):
        root = root.resolve()
        entrada = _existing_layout_dir(root, "entrada")
        saida = _existing_layout_dir(root, "saida")
        if not entrada or not saida:
            continue
        input_names = [compact(path.stem) for path in entrada.iterdir() if path.is_file()]
        output_names = [compact(path.stem) for path in saida.iterdir() if path.is_file()]
        score = 100 - min(distance, 20)
        if any(name.startswith("contasapagar") and "planodecontas" not in name for name in input_names):
            score += 50
        if any("contasapagar" in name and "planodecontas" in name for name in input_names):
            score += 40
        if any(name.startswith("modeloimportacaocontaspagar") for name in input_names):
            score += 30
        if any(re.fullmatch(r"depara\d*", name) for name in input_names):
            score += 25
        if any(name.startswith("planilhatratadafornecedor") for name in output_names):
            score += 15
        candidates.append((score, -distance, root, entrada, saida))

    if not candidates:
        raise FileNotFoundError(
            "Nao encontrei uma raiz existente contendo simultaneamente as pastas "
            "entrada e saida (ou saída), a partir da pasta do script e seus pais."
        )
    _, _, PROJECT_ROOT, INPUT_DIR, OUTPUT_DIR = max(candidates, key=lambda item: (item[0], item[1]))
    return PROJECT_ROOT, INPUT_DIR, OUTPUT_DIR


def newest(paths: Sequence[Path]) -> Path:
    return max(paths, key=lambda path: path.stat().st_mtime)


def resolve_argument(path: Path | None, default_directory: Path) -> Path | None:
    if path is None:
        return None
    expanded = path.expanduser()
    if expanded.is_absolute():
        return expanded.resolve()
    if expanded.parent == Path("."):
        return (default_directory / expanded).resolve()
    return (PROJECT_ROOT / expanded).resolve()


def resolve_variant(
    stem: str,
    suffixes: set[str],
    directory: Path,
    *,
    required: bool = True,
) -> Path | None:
    wanted = compact(stem)
    exact: list[Path] = []
    variants: list[Path] = []
    for path in directory.iterdir():
        if not path.is_file() or path.name.startswith("~$") or path.suffix.lower() not in suffixes:
            continue
        key = compact(path.stem)
        if key == wanted:
            exact.append(path)
        elif key.startswith(wanted):
            variants.append(path)
    candidates = exact or variants
    if candidates:
        return newest(candidates).resolve()
    if required:
        raise FileNotFoundError(f"Arquivo {stem} nao encontrado em {directory}")
    return None


def resolve_depara(directory: Path) -> Path:
    candidates = [
        path
        for path in directory.iterdir()
        if path.is_file()
        and not path.name.startswith("~$")
        and path.suffix.lower() in {".xlsx", ".xls"}
        and re.fullmatch(r"depara\d*", compact(path.stem))
    ]
    if not candidates:
        raise FileNotFoundError(f"Arquivo unico DE-PARA nao encontrado em {directory}")
    exact = [path for path in candidates if compact(path.stem) == "depara"]
    return newest(exact or candidates).resolve()


def resolve_supplier_primary(directory: Path) -> Path | None:
    accepted = {"planilhatratadafornecedor", "planilhatratadafornecedores"}
    candidates = [
        path
        for path in directory.iterdir()
        if path.is_file()
        and path.suffix.lower() in {".csv", ".xlsx"}
        and compact(path.stem) in accepted
    ]
    return newest(candidates).resolve() if candidates else None


# ---------------------------------------------------------------------------
# Leitura de CSV e XLSX, inclusive XLSX com estilos corrompidos
# ---------------------------------------------------------------------------

def read_csv_table(path: Path) -> Table:
    raw = path.read_bytes()
    decoded: str | None = None
    encoding = "utf-8-sig"
    for candidate in ("utf-8-sig", "cp1252", "latin1"):
        try:
            decoded = raw.decode(candidate)
            encoding = candidate
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise ValueError(f"Nao foi possivel identificar a codificacao de {path}")

    first_line = next((line for line in decoded.splitlines() if line.strip()), "")
    if ";" in first_line:
        delimiter = ";"
    else:
        try:
            delimiter = csv.Sniffer().sniff(decoded[:20000], delimiters=",\t|").delimiter
        except csv.Error:
            delimiter = ","

    all_rows = list(csv.reader(io.StringIO(decoded), delimiter=delimiter))
    if not all_rows:
        raise ValueError(f"Arquivo vazio: {path}")
    headers = list(all_rows[0])
    rows = [
        (line, list(row))
        for line, row in enumerate(all_rows[1:], start=2)
        if any(clean_text(value) for value in row)
    ]
    return Table(
        path=path,
        sheet=path.name,
        header_row=1,
        headers=headers,
        rows=rows,
        delimiter=delimiter,
        encoding=encoding,
    )


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _column_index(reference: str) -> int:
    match = re.match(r"([A-Z]+)", reference.upper())
    if not match:
        return 0
    result = 0
    for character in match.group(1):
        result = result * 26 + (ord(character) - 64)
    return result - 1


def _read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    result: list[str] = []
    for item in root.iter():
        if _local_name(item.tag) != "si":
            continue
        result.append("".join(node.text or "" for node in item.iter() if _local_name(node.tag) == "t"))
    return result


def _xlsx_sheet_targets(archive: zipfile.ZipFile) -> tuple[list[tuple[str, str]], bool]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    date1904 = False
    for node in workbook.iter():
        if _local_name(node.tag) == "workbookPr":
            raw = node.attrib.get("date1904", "0").casefold()
            date1904 = raw in {"1", "true"}
            break

    relationships: dict[str, str] = {}
    rel_path = "xl/_rels/workbook.xml.rels"
    if rel_path in archive.namelist():
        rel_root = ET.fromstring(archive.read(rel_path))
        for node in rel_root:
            relationship_id = node.attrib.get("Id", "")
            target = node.attrib.get("Target", "")
            if relationship_id and target:
                if target.startswith("/"):
                    target = target.lstrip("/")
                elif not target.startswith("xl/"):
                    target = "xl/" + target.lstrip("./")
                relationships[relationship_id] = target

    sheets: list[tuple[str, str]] = []
    for node in workbook.iter():
        if _local_name(node.tag) != "sheet":
            continue
        name = node.attrib.get("name", "Sheet")
        relationship_id = next((value for key, value in node.attrib.items() if key.endswith("}id") or key == "r:id"), "")
        target = relationships.get(relationship_id, "")
        if target and target in archive.namelist():
            sheets.append((name, target))
    return sheets, date1904


def _read_xlsx_sheet_xml(
    archive: zipfile.ZipFile,
    target: str,
    shared_strings: Sequence[str],
) -> list[tuple[int, list[Any]]]:
    row_cells: dict[int, dict[int, Any]] = defaultdict(dict)
    max_column = -1
    with archive.open(target) as handle:
        for _, element in ET.iterparse(handle, events=("end",)):
            if _local_name(element.tag) != "c":
                continue
            reference = element.attrib.get("r", "")
            row_match = re.search(r"\d+", reference)
            if not row_match:
                element.clear()
                continue
            row_number = int(row_match.group())
            column = _column_index(reference)
            cell_type = element.attrib.get("t", "")
            value_node = next((child for child in element if _local_name(child.tag) == "v"), None)
            inline_node = next((child for child in element if _local_name(child.tag) == "is"), None)
            value: Any = ""
            if cell_type == "inlineStr" and inline_node is not None:
                value = "".join(node.text or "" for node in inline_node.iter() if _local_name(node.tag) == "t")
            elif value_node is not None and value_node.text is not None:
                raw = value_node.text
                if cell_type == "s":
                    try:
                        index = int(raw)
                        value = shared_strings[index] if 0 <= index < len(shared_strings) else ""
                    except ValueError:
                        value = ""
                elif cell_type == "b":
                    value = raw == "1"
                elif cell_type in {"str", "e"}:
                    value = raw
                else:
                    try:
                        numeric = float(raw)
                        value = int(numeric) if numeric.is_integer() else numeric
                    except ValueError:
                        value = raw
            row_cells[row_number][column] = value
            max_column = max(max_column, column)
            element.clear()

    rows: list[tuple[int, list[Any]]] = []
    width = max_column + 1
    for row_number in sorted(row_cells):
        values = [row_cells[row_number].get(column, "") for column in range(width)]
        if any(clean_text(value) for value in values):
            rows.append((row_number, values))
    return rows


def read_xlsx_sheets_xml(path: Path) -> list[RawSheet]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = _read_shared_strings(archive)
        targets, date1904 = _xlsx_sheet_targets(archive)
        if not targets:
            raise ValueError(f"Nenhuma planilha encontrada em {path}")
        return [
            RawSheet(
                path=path,
                name=name,
                rows=_read_xlsx_sheet_xml(archive, target, shared_strings),
                date1904=date1904,
            )
            for name, target in targets
        ]


def read_xlsx_sheets(path: Path) -> list[RawSheet]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return read_xlsx_sheets_xml(path)

    try:
        date1904 = bool(getattr(workbook, "epoch", None) and getattr(workbook.epoch, "year", 1899) == 1904)
        result: list[RawSheet] = []
        for worksheet in workbook.worksheets:
            rows: list[tuple[int, list[Any]]] = []
            for row_number, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
                row = list(values)
                if any(clean_text(value) for value in row):
                    rows.append((row_number, row))
            result.append(RawSheet(path=path, name=worksheet.title, rows=rows, date1904=date1904))
        return result
    finally:
        workbook.close()


def find_header(headers: Sequence[Any], aliases: Sequence[str], *, required: bool = True) -> int | None:
    normalized_headers = [norm(header) for header in headers]
    compact_headers = [compact(header) for header in headers]
    normalized_aliases = [norm(alias) for alias in aliases]
    compact_aliases = [compact(alias) for alias in aliases]

    for alias in normalized_aliases:
        for index, header in enumerate(normalized_headers):
            if alias and header == alias:
                return index
    for alias in compact_aliases:
        for index, header in enumerate(compact_headers):
            if alias and header == alias:
                return index

    candidates: list[tuple[int, int]] = []
    for alias in normalized_aliases:
        alias_tokens = set(alias.split())
        if not alias_tokens:
            continue
        for index, header in enumerate(normalized_headers):
            header_tokens = set(header.split())
            if alias_tokens.issubset(header_tokens):
                candidates.append((len(alias_tokens), index))
    if candidates:
        return max(candidates, key=lambda item: (item[0], -item[1]))[1]

    if required:
        raise KeyError(f"Coluna nao encontrada. Esperado um de: {', '.join(aliases)}")
    return None


def locate_xlsx_table(
    path: Path,
    signature: dict[str, Sequence[str]],
    *,
    minimum_score: int,
    preferred_sheets: Sequence[str] = (),
) -> tuple[Table, int]:
    if path.suffix.lower() == ".xls":
        raise ValueError(
            f"{path.name} esta em formato .xls. Salve uma copia como .xlsx sem alterar as abas."
        )
    sheets = read_xlsx_sheets(path)
    preferred = {norm(name) for name in preferred_sheets}
    best: tuple[int, int, RawSheet, int, list[Any]] | None = None
    for sheet in sheets:
        sheet_bonus = 5 if norm(sheet.name) in preferred else 0
        for row_number, row in sheet.rows[:30]:
            score = sum(find_header(row, aliases, required=False) is not None for aliases in signature.values())
            candidate = (score, sheet_bonus, sheet, row_number, row)
            if best is None or (score, sheet_bonus) > (best[0], best[1]):
                best = candidate
    if best is None or best[0] < minimum_score:
        raise ValueError(
            f"Nao identifiquei o layout esperado em {path.name}. "
            f"Pontuacao encontrada: {best[0] if best else 0}; minima: {minimum_score}."
        )
    score, _, sheet, header_row, headers = best
    rows = [
        (row_number, list(row))
        for row_number, row in sheet.rows
        if row_number > header_row and any(clean_text(value) for value in row)
    ]
    return (
        Table(
            path=path,
            sheet=sheet.name,
            header_row=header_row,
            headers=list(headers),
            rows=rows,
            date1904=sheet.date1904,
        ),
        score,
    )


EXTRACTION_SIGNATURE: dict[str, Sequence[str]] = {
    "id": ["Id"],
    "supplier": ["Favorecido"],
    "franchise": ["Franquia"],
    "account_type": ["Tipo de Conta"],
    "document": ["Documento"],
    "cost_center": ["Centro de Custo"],
    "status": ["Status"],
    "amount": ["ValorTotal", "Valor Total"],
    "due": ["Data Vencimento", "Data de Vencimento"],
    "paid": ["Data Pagamento", "Data de Pagamento"],
}

ACCOUNT_SIGNATURE: dict[str, Sequence[str]] = {
    "id": ["Id"],
    "account_code": ["Cod Conta", "Codigo Conta", "Codigo da Conta"],
    "account_name": ["Nome da Conta"],
}

SUPPLIER_SIGNATURE: dict[str, Sequence[str]] = {
    "code": ["Codigo", "Codigo do Fornecedor"],
    "fantasy": ["Nome Fantasia", "Nome"],
    "legal": ["Razao Social"],
}


def discover_extraction(directory: Path, explicit: Path | None = None) -> Path:
    if explicit is not None:
        if not explicit.is_file():
            raise FileNotFoundError(explicit)
        locate_xlsx_table(explicit, EXTRACTION_SIGNATURE, minimum_score=7)
        return explicit.resolve()

    matches: list[tuple[int, int, float, Path]] = []
    for path in directory.iterdir():
        key = compact(path.stem)
        if (
            not path.is_file()
            or path.name.startswith("~$")
            or path.suffix.lower() != ".xlsx"
            or "planodecontas" in key
            or key.startswith("modeloimportacao")
            or re.fullmatch(r"depara\d*", key)
        ):
            continue
        try:
            table, score = locate_xlsx_table(path, EXTRACTION_SIGNATURE, minimum_score=7)
        except Exception:
            continue
        matches.append((score, len(table.rows), path.stat().st_mtime, path))
    if not matches:
        raise FileNotFoundError(f"Extracao de Contas a Pagar nao encontrada em {directory}")
    return max(matches, key=lambda item: (item[0], item[1], item[2]))[3].resolve()


def discover_account_relation(directory: Path, explicit: Path | None = None) -> Path:
    if explicit is not None:
        if not explicit.is_file():
            raise FileNotFoundError(explicit)
        locate_xlsx_table(explicit, ACCOUNT_SIGNATURE, minimum_score=2)
        return explicit.resolve()

    preferred = [
        path
        for path in directory.iterdir()
        if path.is_file()
        and path.suffix.lower() == ".xlsx"
        and "contasapagar" in compact(path.stem)
        and "planodecontas" in compact(path.stem)
    ]
    candidates = preferred or [
        path
        for path in directory.iterdir()
        if path.is_file() and path.suffix.lower() == ".xlsx" and "planodecontas" in compact(path.stem)
    ]
    matches: list[tuple[int, int, float, Path]] = []
    for path in candidates:
        try:
            table, score = locate_xlsx_table(path, ACCOUNT_SIGNATURE, minimum_score=2)
        except Exception:
            continue
        matches.append((score, len(table.rows), path.stat().st_mtime, path))
    if not matches:
        raise FileNotFoundError(f"Relacao de plano de contas nao encontrada em {directory}")
    return max(matches, key=lambda item: (item[0], item[1], item[2]))[3].resolve()


def read_supplier_table(path: Path) -> Table:
    if path.suffix.lower() == ".csv":
        return read_csv_table(path)
    table, _ = locate_xlsx_table(path, SUPPLIER_SIGNATURE, minimum_score=2)
    return table


# ---------------------------------------------------------------------------
# Relacionamentos: plano de contas, forma de pagamento e fornecedores
# ---------------------------------------------------------------------------

def row_value(row: Sequence[Any], index: int | None) -> Any:
    if index is None or index < 0 or index >= len(row):
        return ""
    return row[index]


def load_account_relations(path: Path, report: ValidationReport) -> dict[str, AccountRelation]:
    table, _ = locate_xlsx_table(path, ACCOUNT_SIGNATURE, minimum_score=2)
    id_index = find_header(table.headers, ACCOUNT_SIGNATURE["id"])
    code_index = find_header(table.headers, ACCOUNT_SIGNATURE["account_code"])
    name_index = find_header(table.headers, ACCOUNT_SIGNATURE["account_name"], required=False)
    relations: dict[str, AccountRelation] = {}

    for line, row in table.rows:
        source_id = normalize_id(row_value(row, id_index))
        raw_code = row_value(row, code_index)
        code = plan_account_code(raw_code)
        name = clean_text(row_value(row, name_index))
        if not source_id:
            if any(clean_text(value) for value in row):
                report.error(
                    "RELACAO_CONTA_SEM_ID",
                    file=path,
                    line=line,
                    field=table.headers[id_index],
                    detail="Linha preenchida sem Id para relacionar o plano de contas.",
                )
            continue
        if not code:
            report.error(
                "CODIGO_CONTA_INVALIDO",
                file=path,
                line=line,
                source_id=source_id,
                field=table.headers[code_index],
                origin=raw_code,
                detail="Nenhum numero permaneceu apos a limpeza do codigo do plano de contas.",
            )
            continue
        relation = AccountRelation(code=code, name=name, source_line=line)
        previous = relations.get(source_id)
        if previous is not None and (previous.code != relation.code or previous.name != relation.name):
            report.error(
                "RELACAO_CONTA_DUPLICADA_CONFLITANTE",
                file=path,
                line=line,
                source_id=source_id,
                origin=f"{previous.code} | {previous.name}",
                compared=f"{relation.code} | {relation.name}",
                detail="O mesmo Id possui mais de uma relacao de plano de contas.",
            )
            continue
        relations[source_id] = relation

    if not relations:
        raise ValueError(f"Nenhuma relacao valida de plano de contas encontrada em {path.name}")
    return relations


def _select_fp_sheet(path: Path) -> RawSheet:
    if path.suffix.lower() == ".xls":
        raise ValueError(
            "O DE-PARA esta em .xls. Salve uma copia como DE-PARA.xlsx, preservando suas abas."
        )
    sheets = read_xlsx_sheets(path)
    aliases = {
        "fp",
        "forma de pagamento",
        "formas de pagamento",
        "formas pagamento",
        "pagamento",
    }
    normalized = {norm(alias) for alias in aliases}
    exact = [sheet for sheet in sheets if norm(sheet.name) in normalized]
    if exact:
        return exact[0]
    partial = [
        sheet
        for sheet in sheets
        if any(alias in norm(sheet.name) or norm(sheet.name) in alias for alias in normalized)
    ]
    if partial:
        return partial[0]
    raise KeyError(
        f"Aba FP/Formas de Pagamento nao encontrada em {path.name}. "
        f"Abas: {', '.join(sheet.name for sheet in sheets)}"
    )


PAYMENT_SOURCE_ALIASES: list[str] = [
    "Forma de Pagamento",
    "Forma Pagamento",
    "Forma Pgto",
    "Forma Pg",
    "Forma",
    "Nome",
    "Descricao FP",
    "Descricao",
    "Descricao da Forma de Pagamento",
    "Nome da Forma de Pagamento",
    "Nome Forma Pagamento",
    "Origem",
    "Forma Origem",
    "Pagamento Origem",
    "De",
    "Antigo",
    "Sistema Antigo",
    "Cadastro Antigo",
    "Cadastro de Origem",
    "Sistema de Origem",
    "Descricao Origem",
    "Nome Origem",
]

PAYMENT_CODE_ALIASES: list[str] = [
    "Codigo Forma de Pagamento",
    "Codigo da Forma de Pagamento",
    "Cod Forma de Pagamento",
    "Cod da Forma de Pagamento",
    "Codigo FP",
    "Cod FP",
    "Codigo",
    "Cod",
    "Id",
    "Destino",
    "Forma Destino",
    "Pagamento Destino",
    "Para",
    "Novo",
    "Sistema Novo",
    "Cadastro Novo",
    "Cadastro de Destino",
    "Sistema de Destino",
    "Descricao Destino",
    "Nome Destino",
]

PAYMENT_ACCOUNT_ALIASES: list[str] = [
    "Codigo da Conta Origem",
    "Codigo Conta Origem",
    "Cod Conta Origem",
    "Conta Origem",
    "Codigo da Conta Destino",
    "Codigo Conta Destino",
    "Cod Conta Destino",
    "Conta Destino",
    "Conta (uso financeiro)",
    "Conta uso financeiro",
    "Codigo da Conta (uso financeiro)",
    "Codigo Conta uso financeiro",
    "Conta Financeira",
    "Conta Financeiro",
]


def _payment_header_table(sheet: RawSheet) -> Table:
    """Localiza a tabela da aba FP sem exigir nomes fixos de cabecalho.

    Alguns arquivos DE-PARA usam cabecalhos como ``De/Para``, ``Antigo/Novo``
    ou apenas os nomes dos sistemas. Primeiro tentamos reconhecer os nomes; se
    isso nao for possivel, usamos o primeiro conjunto de duas ou mais colunas
    preenchidas que possua linhas de dados logo abaixo. Esse fallback replica o
    comportamento ja usado no script de Contas a Receber.
    """
    if not sheet.rows:
        raise ValueError(f"A aba {sheet.name} do DE-PARA esta vazia.")

    recognized: list[tuple[int, int, list[Any]]] = []
    fallback: list[tuple[int, int, list[Any]]] = []
    inspected = sheet.rows[:50]

    for position, (row_number, row) in enumerate(inspected):
        populated = [index for index, value in enumerate(row) if clean_text(value)]
        if len(populated) < 2:
            continue

        source_index = find_header(row, PAYMENT_SOURCE_ALIASES, required=False)
        code_index = find_header(row, PAYMENT_CODE_ALIASES, required=False)
        alias_score = int(source_index is not None) + int(code_index is not None)
        if alias_score == 2 and source_index != code_index:
            recognized.append((row_number, position, list(row)))
            continue

        # Confirma que ha pelo menos uma linha posterior preenchida nas mesmas
        # colunas. Isso evita escolher uma linha de titulo com duas celulas.
        support = 0
        for _, next_row in sheet.rows[position + 1 : position + 11]:
            if sum(bool(clean_text(row_value(next_row, index))) for index in populated) >= 2:
                support += 1
        if support:
            fallback.append((row_number, position, list(row)))

    if recognized:
        header_row, _, headers = min(recognized, key=lambda item: item[0])
    elif fallback:
        header_row, _, headers = min(fallback, key=lambda item: item[0])
    else:
        candidates = [
            (row_number, list(row))
            for row_number, row in inspected
            if len([value for value in row if clean_text(value)]) >= 2
        ]
        if not candidates:
            preview = " | ".join(
                f"linha {row_number}: {', '.join(clean_text(value) for value in row if clean_text(value))}"
                for row_number, row in inspected[:5]
            )
            raise ValueError(
                f"A aba {sheet.name} do DE-PARA nao possui ao menos duas colunas preenchidas. "
                f"Primeiras linhas: {preview or 'sem conteudo'}"
            )
        header_row, headers = candidates[0]

    rows = [
        (row_number, list(row))
        for row_number, row in sheet.rows
        if row_number > header_row and any(clean_text(value) for value in row)
    ]
    return Table(
        path=sheet.path,
        sheet=sheet.name,
        header_row=header_row,
        headers=list(headers),
        rows=rows,
        date1904=sheet.date1904,
    )


def _payment_column_indexes(table: Table) -> tuple[int, int, int | None]:
    source_index = find_header(table.headers, PAYMENT_SOURCE_ALIASES, required=False)
    code_index = find_header(table.headers, PAYMENT_CODE_ALIASES, required=False)
    account_index = find_header(table.headers, PAYMENT_ACCOUNT_ALIASES, required=False)

    # Aliases genericos como Codigo/Origem podem coincidir com a coluna
    # especifica de conta. A conta nunca pode ser usada como forma ou codigo.
    if source_index == account_index:
        source_index = None
    if code_index == account_index:
        code_index = None
    if source_index is not None and source_index == code_index:
        code_index = None

    populated = [index for index, header in enumerate(table.headers) if clean_text(header)]
    if len(populated) < 2:
        raise ValueError(
            f"A aba {table.sheet} do DE-PARA nao possui duas colunas utilizaveis no cabecalho "
            f"da linha {table.header_row}."
        )

    # Quando os nomes nao sao reconhecidos, preserva a convencao do DE-PARA:
    # primeira coluna preenchida = descricao/origem; segunda = codigo/destino.
    if source_index is None:
        source_index = next(
            (index for index in populated if index != code_index and index != account_index),
            populated[0],
        )

    if code_index is None or code_index == source_index:
        code_index = next(
            (
                index
                for index in populated
                if index != source_index and index != account_index
            ),
            -1,
        )
    if code_index < 0:
        # Um cabecalho generico como "Conta" pode ter sido confundido com a
        # coluna opcional de conta. Nesse caso, a segunda coluna continua sendo
        # o codigo da forma de pagamento e a conta usa o fallback do script.
        code_index = next((index for index in populated if index != source_index), -1)
        if code_index == account_index:
            account_index = None

    if code_index < 0 or code_index == source_index:
        raise ValueError(
            f"Nao foi possivel separar descricao e codigo na aba {table.sheet}, "
            f"cabecalho da linha {table.header_row}: "
            + " | ".join(clean_text(value) for value in table.headers if clean_text(value))
        )

    # Aceita uma terceira coluna chamada simplesmente Conta, desde que ela nao
    # seja a coluna de descricao nem a de codigo.
    if account_index is None and len(populated) >= 3:
        for index in populated:
            if index in {source_index, code_index}:
                continue
            header = norm(table.headers[index])
            if header == "conta" or "conta origem" in header or "conta destino" in header:
                account_index = index
                break

    if account_index in {source_index, code_index}:
        account_index = None
    return source_index, code_index, account_index



def _payment_cell_match(cell: Any, expected: str) -> tuple[int, int]:
    """Pontua a correspondencia entre uma celula da aba FP e um nome esperado."""
    cell_key = norm(cell)
    expected_key = norm(expected)
    if not cell_key or not expected_key:
        return (0, 0)
    if cell_key == expected_key:
        return (4, len(expected_key))
    if compact(cell_key) == compact(expected_key):
        return (3, len(expected_key))
    cell_tokens = set(cell_key.split())
    expected_tokens = set(expected_key.split())
    if expected_tokens and expected_tokens.issubset(cell_tokens):
        return (2, len(expected_tokens))
    if cell_tokens and cell_tokens.issubset(expected_tokens):
        return (1, len(cell_tokens))
    return (0, 0)


def _looks_like_account_code(value: Any) -> bool:
    text = clean_text(value)
    code = plan_account_code(text)
    if not code or len(code) < 3:
        return False
    return any(separator in code for separator in ".-/:,") and sum(ch.isdigit() for ch in code) >= 3


def _scan_expected_payment_rules(
    sheet: RawSheet,
    expected_sources: Sequence[str],
) -> dict[str, PaymentRule]:
    """Localiza somente as formas realmente usadas, mesmo sem cabecalho reconhecivel.

    O fallback procura o nome esperado em qualquer celula da aba FP. Depois usa
    as outras celulas preenchidas da mesma linha para identificar o codigo e,
    quando houver, a conta de origem. Isso cobre layouts sem cabecalho, com
    cabecalho mesclado, invertidos ou com nomes de colunas particulares do
    cliente, sem copiar dados de arquivos modelo.
    """
    expected_unique: list[str] = []
    seen_expected: set[str] = set()
    for value in expected_sources:
        cleaned = clean_text(value)
        key = norm(cleaned)
        if cleaned and key and key not in seen_expected:
            expected_unique.append(cleaned)
            seen_expected.add(key)

    rules: dict[str, PaymentRule] = {}
    for expected in expected_unique:
        matches: list[tuple[int, int, int, int, list[Any]]] = []
        for row_number, row in sheet.rows:
            for index, value in enumerate(row):
                strength, size = _payment_cell_match(value, expected)
                if strength:
                    # Prefer exact/long matches and, in ties, the earliest row.
                    matches.append((strength, size, -row_number, index, list(row)))
        if not matches:
            continue

        _, _, neg_row_number, source_index, row = max(matches)
        row_number = -neg_row_number
        populated = [
            index
            for index, value in enumerate(row)
            if index != source_index and clean_text(value)
        ]
        if not populated:
            continue

        account_candidates = [index for index in populated if _looks_like_account_code(row_value(row, index))]
        non_account_candidates = [index for index in populated if index not in account_candidates]

        # Normalmente o codigo esta na celula imediatamente a direita. Quando
        # existe uma conta contábil pontuada na mesma linha, ela e separada do
        # codigo da forma de pagamento.
        def code_score(index: int) -> tuple[int, int, int]:
            distance = abs(index - source_index)
            side_bonus = 20 if index > source_index else 10
            content = clean_text(row_value(row, index))
            digit_bonus = 3 if any(ch.isdigit() for ch in content) else 0
            return (side_bonus - distance, digit_bonus, -index)

        candidate_pool = non_account_candidates or populated
        code_index = max(candidate_pool, key=code_score)
        code = clean_text(row_value(row, code_index))
        if not code:
            continue

        account_origin = ""
        remaining_accounts = [index for index in account_candidates if index != code_index]
        if remaining_accounts:
            account_index = min(remaining_accounts, key=lambda index: (abs(index - code_index), index))
            account_origin = plan_account_code(row_value(row, account_index))

        source_name = clean_text(row_value(row, source_index)) or expected
        rules[norm(expected)] = PaymentRule(
            source_name=source_name,
            code=code,
            account_origin=account_origin,
        )

    return rules


def _payment_sheet_preview(sheet: RawSheet, limit: int = 12) -> str:
    parts: list[str] = []
    for row_number, row in sheet.rows[:limit]:
        values = [clean_text(value) for value in row if clean_text(value)]
        if values:
            parts.append(f"L{row_number}: " + " | ".join(values[:8]))
    return " ; ".join(parts) or "aba sem conteudo"


def load_payment_rules(
    path: Path,
    expected_sources: Sequence[str] = (),
) -> dict[str, PaymentRule]:
    """Carrega a aba FP sem tornar um nome de cabecalho obrigatorio.

    A leitura direta das formas efetivamente usadas e executada sempre. A
    tabela por colunas e apenas uma segunda fonte: quando o cabecalho e
    reconhecido ela prevalece; quando foi inferido, a leitura direta prevalece.
    """
    sheet = _select_fp_sheet(path)
    expected_keys = {norm(value) for value in expected_sources if norm(value)}
    scanned = _scan_expected_payment_rules(sheet, expected_sources)

    rules: dict[str, PaymentRule] = {}
    table: Table | None = None
    header_mode = "nao identificado"
    header_error = ""
    recognized_header = False

    try:
        table = _payment_header_table(sheet)
        recognized_source = find_header(
            table.headers,
            PAYMENT_SOURCE_ALIASES,
            required=False,
        )
        recognized_code = find_header(
            table.headers,
            PAYMENT_CODE_ALIASES,
            required=False,
        )
        recognized_header = (
            recognized_source is not None
            and recognized_code is not None
            and recognized_source != recognized_code
        )
        header_mode = "reconhecido" if recognized_header else "heuristico"
        source_index, code_index, account_index = _payment_column_indexes(table)

        for _, row in table.rows:
            source_name = clean_text(row_value(row, source_index))
            code = clean_text(row_value(row, code_index))
            account_origin = (
                plan_account_code(row_value(row, account_index))
                if account_index is not None
                else ""
            )
            if not source_name or not code:
                continue

            key = norm(source_name)
            if not key:
                continue
            rule = PaymentRule(
                source_name=source_name,
                code=code,
                account_origin=account_origin,
            )
            previous = rules.get(key)
            if previous is not None and previous != rule:
                if recognized_header:
                    raise ValueError(
                        f"Aba {table.sheet} possui relacao conflitante para "
                        f"{source_name}: {previous.code} e {rule.code}."
                    )
                # Em cabecalho apenas inferido, uma coluna errada pode gerar
                # conflito artificial. Remove a chave e deixa a busca direta
                # das formas necessarias decidir com base na linha completa.
                rules.pop(key, None)
                continue
            rules[key] = rule
    except Exception as exc:
        if isinstance(exc, ValueError) and recognized_header and "conflitante" in str(exc):
            raise
        header_error = f"{type(exc).__name__}: {exc}"
        table = None
        rules = {}
        header_mode = "nao identificado"

    # Para cabecalho reconhecido, a tabela e a fonte principal. Para cabecalho
    # apenas inferido ou ausente, a busca direta substitui a relacao da forma
    # efetivamente usada, evitando tratar uma coluna contabil como codigo FP.
    for key, rule in scanned.items():
        if key not in rules or not recognized_header:
            rules[key] = rule

    if not rules:
        details = (
            f" Motivo da leitura tabular: {header_error}."
            if header_error
            else ""
        )
        raise ValueError(
            f"Nao encontrei relacoes utilizaveis na aba {sheet.name} do DE-PARA. "
            f"O leitor pesquisou diretamente as formas necessarias e tambem "
            f"tentou inferir as colunas sem exigir nomes fixos.{details} "
            f"Amostra: {_payment_sheet_preview(sheet)}"
        )

    resolved_expected = sum(
        1
        for key in expected_keys
        if payment_lookup(rules, key)[0] is not None
    )
    table_line = table.header_row if table is not None else "-"
    diagnostic = (
        f"DE-PARA FP: aba={sheet.name!r} | cabecalho={header_mode} "
        f"| linha={table_line} | regras={len(rules)} "
        f"| formas_necessarias={len(expected_keys)} "
        f"| formas_localizadas={resolved_expected}"
    )
    if header_error:
        diagnostic += f" | leitura_tabular={header_error}"
    print(diagnostic)
    return rules

def payment_lookup(rules: dict[str, PaymentRule], value: Any) -> tuple[PaymentRule | None, str]:
    """Localiza uma regra pelo nome de origem ou pelo valor de destino.

    Aceitar o destino e importante quando a extracao nao possui forma de
    pagamento e o parametro padrao foi informado exatamente como o valor que
    deve ser gravado. Mesmo nesse caso, o valor continua vindo da aba FP do
    DE-PARA; o parametro apenas escolhe uma relacao existente.
    """
    key = norm(value)
    if not key:
        return None, "VAZIO"
    if key in rules:
        return rules[key], "ORIGEM_EXATA"

    # O nome informado tambem pode ser o codigo/descricao de destino da aba FP.
    destination_candidates = {
        (rule.code, rule.account_origin, rule.source_name): rule
        for rule in rules.values()
        if norm(rule.code) == key or compact(rule.code) == compact(value)
    }
    if len(destination_candidates) == 1:
        return next(iter(destination_candidates.values())), "DESTINO_EXATO"

    key_tokens = set(key.split())
    candidates: list[PaymentRule] = []
    for rule_key, rule in rules.items():
        rule_tokens = set(rule_key.split())
        if key_tokens and (key_tokens.issubset(rule_tokens) or rule_tokens.issubset(key_tokens)):
            candidates.append(rule)
    unique = {(candidate.code, candidate.account_origin, candidate.source_name): candidate for candidate in candidates}
    if len(unique) == 1:
        return next(iter(unique.values())), "CONTENCAO_SEGURA"

    families = {
        "credito": {"credito", "cartao"},
        "debito": {"debito", "cartao"},
        "dinheiro": {"dinheiro"},
        "pix": {"pix"},
        "financeiro": {"conta", "financeiro"},
    }
    for family, tokens in families.items():
        if family in key or tokens.issubset(key_tokens):
            family_candidates = [
                rule
                for rule_key, rule in rules.items()
                if family in rule_key or tokens.issubset(set(rule_key.split()))
            ]
            unique_family = {
                (candidate.code, candidate.account_origin, candidate.source_name): candidate
                for candidate in family_candidates
            }
            if len(unique_family) == 1:
                return next(iter(unique_family.values())), "FAMILIA_SEGURA"
    return None, "NAO_LOCALIZADO"



def _unique_payment_rules(rules: dict[str, PaymentRule]) -> list[PaymentRule]:
    unique: dict[tuple[str, str, str], PaymentRule] = {}
    for rule in rules.values():
        key = (
            clean_text(rule.source_name),
            clean_text(rule.code),
            plan_account_code(rule.account_origin),
        )
        unique[key] = rule
    return sorted(
        unique.values(),
        key=lambda rule: (norm(rule.source_name), clean_text(rule.code), rule.account_origin),
    )


def _payment_rules_text(
    rules: dict[str, PaymentRule],
    *,
    default_account_origin: str = "",
) -> str:
    fallback = plan_account_code(default_account_origin)
    parts: list[str] = []
    for rule in _unique_payment_rules(rules):
        account = plan_account_code(rule.account_origin) or fallback or "sem conta"
        parts.append(
            f"{clean_text(rule.source_name)} -> codigo {clean_text(rule.code)} -> conta {account}"
        )
    return " ; ".join(parts) or "nenhuma regra"


def print_payment_rules(
    rules: dict[str, PaymentRule],
    *,
    default_account_origin: str = "",
) -> None:
    print("Formas de pagamento encontradas na aba FP:")
    fallback = plan_account_code(default_account_origin)
    for rule in _unique_payment_rules(rules):
        account = plan_account_code(rule.account_origin) or fallback or "-"
        print(
            f"  - nome={clean_text(rule.source_name)!r} | "
            f"codigo={clean_text(rule.code)!r} | conta_origem={account!r}"
        )


def select_payment_rule(
    rules: dict[str, PaymentRule],
    selector: Any,
) -> tuple[PaymentRule | None, str]:
    """Resolve uma regra por nome ou pelo codigo presente na aba FP."""
    cleaned = clean_text(selector)
    if not cleaned:
        return None, "NAO_INFORMADO"

    by_name, method = payment_lookup(rules, cleaned)
    if by_name is not None:
        return by_name, method

    code_matches = [
        rule
        for rule in _unique_payment_rules(rules)
        if clean_text(rule.code).casefold() == cleaned.casefold()
    ]
    unique_code_matches = {
        (rule.code, plan_account_code(rule.account_origin), rule.source_name): rule
        for rule in code_matches
    }
    if len(unique_code_matches) == 1:
        return next(iter(unique_code_matches.values())), "CODIGO_EXATO"
    if len(unique_code_matches) > 1:
        return None, "CODIGO_AMBIGUO"
    return None, method


def _prompt_payment_rule(
    rules: dict[str, PaymentRule],
    *,
    default_account_origin: str,
) -> tuple[PaymentRule, str]:
    """Solicita uma escolha explicita quando a extracao nao informa a forma."""
    available = _unique_payment_rules(rules)
    fallback = plan_account_code(default_account_origin)
    print()
    print("A extracao de Contas a Pagar nao informa a Forma de Pagamento.")
    print("Escolha a forma padrao que sera aplicada a todos os lancamentos desta execucao:")
    for position, rule in enumerate(available, start=1):
        account = plan_account_code(rule.account_origin) or fallback or "-"
        print(
            f"  [{position}] {clean_text(rule.source_name)} "
            f"-> codigo {clean_text(rule.code)} -> conta {account}"
        )
    print(
        "Digite o numero, o nome ou o codigo da forma. "
        "Pressione ENTER sem valor para cancelar."
    )

    while True:
        try:
            answer = clean_text(input("Forma de pagamento padrao: "))
        except (EOFError, OSError) as exc:
            raise ValueError(
                "Nao foi possivel ler a escolha da Forma de Pagamento no terminal. "
                "Execute novamente informando --forma-pagamento-padrao seguido do "
                "nome ou codigo existente na aba FP."
            ) from exc

        if not answer:
            raise ValueError(
                "Execucao cancelada porque nenhuma Forma de Pagamento foi escolhida. "
                "Execute novamente e selecione uma opcao, ou informe "
                "--forma-pagamento-padrao."
            )

        if answer.isdigit():
            position = int(answer)
            if 1 <= position <= len(available):
                return available[position - 1], "ESCOLHA_INTERATIVA_NUMERO"

        selected, method = select_payment_rule(rules, answer)
        if selected is not None:
            return selected, f"ESCOLHA_INTERATIVA_{method}"

        print(
            f"Opcao {answer!r} nao localizada. "
            f"Escolha um numero entre 1 e {len(available)}, um nome ou um codigo listado."
        )


def choose_default_payment_rule(
    rules: dict[str, PaymentRule],
    selector: str,
    *,
    default_account_origin: str,
    allow_interactive_prompt: bool,
) -> tuple[PaymentRule, str]:
    """Escolhe o padrao sem inventar uma forma ausente da extracao.

    Quando o usuario informa nome/codigo, a selecao deve existir no DE-PARA.
    Sem seletor, a escolha automatica so ocorre se todas as linhas da aba FP
    resultarem no mesmo codigo e na mesma conta de origem efetiva. Quando ha
    varias alternativas e o terminal e interativo, o usuario escolhe uma delas.
    """
    cleaned_selector = clean_text(selector)
    if cleaned_selector:
        selected, method = select_payment_rule(rules, cleaned_selector)
        if selected is not None:
            return selected, f"PARAMETRO_{method}"
        special = ""
        if norm(cleaned_selector) in {norm(value) for value in PAYMENT_ACCOUNT_ALIASES}:
            special = (
                " Esse texto parece ser o cabecalho da coluna de conta financeira, "
                "nao o nome de uma forma de pagamento."
            )
        raise ValueError(
            "Forma de pagamento padrao nao localizada na aba FP: "
            f"{cleaned_selector!r}. Metodo: {method}.{special} "
            "Regras disponiveis: "
            + _payment_rules_text(
                rules,
                default_account_origin=default_account_origin,
            )
        )

    fallback = plan_account_code(default_account_origin)
    groups: dict[tuple[str, str], PaymentRule] = {}
    for rule in _unique_payment_rules(rules):
        effective_account = plan_account_code(rule.account_origin) or fallback
        groups[(clean_text(rule.code), effective_account)] = rule
    if len(groups) == 1:
        return next(iter(groups.values())), "AUTO_RELACAO_UNICA"

    if allow_interactive_prompt and sys.stdin is not None and sys.stdin.isatty():
        return _prompt_payment_rule(
            rules,
            default_account_origin=default_account_origin,
        )

    raise ValueError(
        "A extracao de Contas a Pagar nao possui coluna de Forma de Pagamento e "
        f"a aba FP contem {len(groups)} relacoes diferentes. Em uma execucao sem "
        "terminal interativo, informe o nome ou o codigo com "
        "--forma-pagamento-padrao. Regras disponiveis: "
        + _payment_rules_text(
            rules,
            default_account_origin=default_account_origin,
        )
    )


def add_supplier_to_catalog(catalog: SupplierCatalog, code: Any, *names: Any) -> None:
    normalized_code = normalize_id(code)
    if not normalized_code:
        return
    if normalized_code.isdigit():
        catalog.used_codes.add(int(normalized_code))
    for name in names:
        cleaned = clean_text(name)
        if not cleaned:
            continue
        catalog.by_exact[exact_name_key(cleaned)].add(normalized_code)
        catalog.by_normalized[normalized_name_key(cleaned)].add(normalized_code)
        catalog.by_singular[singular_name_key(cleaned)].add(normalized_code)
        catalog.by_compact[compact(cleaned)].add(normalized_code)  # FORNECEDOR_IDENTIDADE_COMPACTA_V2026_08_25_1


def load_supplier_catalog(paths: Sequence[Path]) -> SupplierCatalog:
    catalog = SupplierCatalog()
    seen_paths: set[Path] = set()
    for path in paths:
        resolved = path.resolve()
        if resolved in seen_paths or not resolved.is_file():
            continue
        seen_paths.add(resolved)
        table = read_supplier_table(resolved)
        code_index = find_header(table.headers, SUPPLIER_SIGNATURE["code"])
        fantasy_index = find_header(table.headers, SUPPLIER_SIGNATURE["fantasy"], required=False)
        legal_index = find_header(table.headers, SUPPLIER_SIGNATURE["legal"], required=False)
        if fantasy_index is None and legal_index is None:
            raise ValueError(f"Planilha de fornecedores sem Nome Fantasia/Razao Social: {path.name}")
        for _, row in table.rows:
            code = row_value(row, code_index)
            fantasy = row_value(row, fantasy_index)
            legal = row_value(row, legal_index)
            if clean_text(code) and (clean_text(fantasy) or clean_text(legal)):
                add_supplier_to_catalog(catalog, code, fantasy, legal)
    return catalog


def match_supplier(catalog: SupplierCatalog, name: Any) -> tuple[str, str]:
    cleaned = clean_text(name)
    if not cleaned:
        return "", "SEM_NOME"

    exact = catalog.by_exact.get(exact_name_key(cleaned), set())
    if len(exact) == 1:
        return next(iter(exact)), "EXATO"
    if len(exact) > 1:
        return "", "AMBIGUO_EXATO"

    normalized = catalog.by_normalized.get(normalized_name_key(cleaned), set())
    if len(normalized) == 1:
        return next(iter(normalized)), "NORMALIZADO"
    if len(normalized) > 1:
        return "", "AMBIGUO_NORMALIZADO"

    singular = catalog.by_singular.get(singular_name_key(cleaned), set())
    if len(singular) == 1:
        return next(iter(singular)), "SINGULAR_PLURAL"
    if len(singular) > 1:
        return "", "AMBIGUO_SINGULAR_PLURAL"

    compact_matches = catalog.by_compact.get(compact(cleaned), set())
    if len(compact_matches) == 1:
        return next(iter(compact_matches)), "COMPACTO"
    if len(compact_matches) > 1:
        return "", "AMBIGUO_COMPACTO"
    return "", "NAO_LOCALIZADO"


def next_supplier_code(used_codes: set[int]) -> str:
    code = 100000
    while code in used_codes:
        code += 1
    if code > 999999:
        raise ValueError("Nao ha codigo de fornecedor disponivel entre 100000 e 999999.")
    used_codes.add(code)
    return str(code)


# ---------------------------------------------------------------------------
# Modelos e montagem de linhas
# ---------------------------------------------------------------------------

MODEL_FIELDS: dict[str, Sequence[str]] = {
    "code": ["Codigo max 6 digitos", "Codigo (max 6 digitos)"],
    "launch": ["Data Lancamento DD MM AAAA Obrigatorio", "Data Lancamento"],
    "competence": ["Data Competencia DD MM AAAA Obrigatorio", "Data Competencia"],
    "due": ["Data Vencimento DD MM AAAA Obrigatorio se Confirmado", "Data Vencimento"],
    "confirmation": ["Data Confirmacao DD MM AAAA Obrigatorio se Confirmado", "Data Confirmacao"],
    "supplier": ["Fornecedor"],
    "account_origin": ["Codigo da Conta Origem Obrigatorio", "Codigo da Conta Origem"],
    "account_destination": ["Codigo da Conta Destino Obrigatorio", "Codigo da Conta Destino"],
    "payment": ["Codigo Forma de Pagamento Obrigatorio", "Codigo Forma de Pagamento"],
    "amount": ["Valor Titulo Obrigatorio", "Valor Titulo"],
    "interest": ["Valor Juros Multa", "Valor Juros/Multa"],
    "paid_amount": ["Valor Pago Obrigatorio se Confirmado", "Valor Pago"],
    "confirmed": ["Confirmado S N", "Confirmado(S/N)"],
    "observation": ["Observacao"],
}

SUPPLIER_MODEL_FIELDS: dict[str, Sequence[str]] = {
    "code": ["Codigo max 6 digitos", "Codigo (max 6 digitos)"],
    "fantasy": ["Nome Fantasia Obrigatorio", "Nome Fantasia"],
    "legal": ["Razao Social Obrigatorio", "Razao Social"],
    "tax_id": ["CPF CNPJ", "CPF/CNPJ"],
    "observation": ["Observacao"],
}


def map_model_fields(headers: Sequence[Any], fields: dict[str, Sequence[str]]) -> dict[str, int | None]:
    result: dict[str, int | None] = {}
    for field_name, aliases in fields.items():
        required = field_name not in {"interest", "tax_id", "observation"}
        result[field_name] = find_header(headers, aliases, required=required)
    return result


def build_supplier_row(
    headers: Sequence[Any],
    indexes: dict[str, int | None],
    supplier: GeneratedSupplier,
    import_date: str,
) -> list[str]:
    row = ["" for _ in headers]
    row[indexes["code"] or 0] = supplier.code
    if indexes["fantasy"] is not None:
        row[indexes["fantasy"]] = supplier.name
    if indexes["legal"] is not None:
        row[indexes["legal"]] = supplier.name
    if indexes["observation"] is not None:
        row[indexes["observation"]] = f"{SUPPLIER_OBSERVATION_PREFIX} | Importação {import_date}"
    return row


def update_existing_supplier_observations(
    table: Table,
    indexes: dict[str, int | None],
    import_date: str,
) -> int:
    """Atualiza a observacao dos fornecedores para a data desta execucao.

    O arquivo complementar pertence exclusivamente aos fornecedores gerados por
    Contas a Pagar. Por isso, toda linha valida deve receber exatamente o texto
    padrao com a data efetivamente usada pela execucao. A alteracao ocorre em
    memoria e so e publicada depois que todo o processamento termina sem erros.
    """
    observation_index = indexes.get("observation")
    if observation_index is None:
        return 0

    target_observation = (
        f"{SUPPLIER_OBSERVATION_PREFIX} | Importação {import_date}"
    )
    updated = 0
    revised_rows: list[tuple[int, list[Any]]] = []
    supplier_identity_fields = ("code", "fantasy", "legal")

    for line, original_row in table.rows:
        row = list(original_row)
        if len(row) <= observation_index:
            row.extend([""] * (observation_index + 1 - len(row)))

        has_supplier_data = any(
            clean_text(row_value(row, indexes.get(field_name)))
            for field_name in supplier_identity_fields
        )
        if has_supplier_data:
            current = clean_text(row[observation_index])
            if current != target_observation:
                row[observation_index] = target_observation
                updated += 1

        revised_rows.append((line, row))

    table.rows = revised_rows
    return updated


def validate_supplier_row(
    headers: Sequence[Any],
    indexes: dict[str, int | None],
    row: Sequence[Any],
) -> list[str]:
    missing: list[str] = []
    for key in ("code", "fantasy", "legal"):
        index = indexes.get(key)
        if index is None or not clean_text(row_value(row, index)):
            missing.append(clean_text(headers[index]) if index is not None else key)
    code = normalize_id(row_value(row, indexes.get("code")))
    if code and (not code.isdigit() or len(code) > 6):
        missing.append("Codigo numerico com no maximo 6 digitos")
    return missing


def required_output_indexes(headers: Sequence[Any]) -> list[int]:
    result: list[int] = []
    for index, header in enumerate(headers):
        normalized = norm(header)
        if "obrigatorio" in normalized and "obrigatorio se" not in normalized:
            result.append(index)
    return result


def source_columns(headers: Sequence[Any]) -> dict[str, int | None]:
    result: dict[str, int | None] = {}
    for field_name, aliases in EXTRACTION_SIGNATURE.items():
        result[field_name] = find_header(headers, aliases, required=field_name != "paid")
    result["payment"] = find_header(
        headers,
        ["Forma de Pagamento", "Forma Pagamento", "Meio de Pagamento"],
        required=False,
    )
    return result


def status_confirmation(value: Any) -> str:
    normalized = norm(value)
    if normalized in {"paga", "pago", "quitada", "quitado", "paga total"}:
        return "S"
    if normalized in {
        "nao paga",
        "nao pago",
        "pendente",
        "em aberto",
        "aberta",
        "aberto",
        "agendada",
        "agendado",
    }:
        return "N"
    return ""


def build_observation(
    values: dict[str, Any],
    account: AccountRelation,
    *,
    import_date: str,
    original_id: str,
    output_code: str,
    unconfirmed_payment_date: str,
) -> str:
    parts: list[str] = []
    if output_code != original_id:
        parts.append(f"Id origem: {original_id}")
    for label, key in (
        ("Franquia", "franchise"),
        ("Tipo de Conta", "account_type"),
        ("Centro de Custo", "cost_center"),
        ("Documento", "document"),
    ):
        value = clean_text(values.get(key, ""))
        if value:
            parts.append(f"{label}: {value}")
    if account.name:
        parts.append(f"Nome da Conta: {account.name}")
    if unconfirmed_payment_date:
        parts.append(
            "Data de pagamento informada na origem apesar de a conta nao estar confirmada: "
            + unconfirmed_payment_date
        )
    parts.append(f"Importação {import_date}")
    return " | ".join(parts)


def assign_transaction_codes(
    source_ids: Sequence[str],
    report: ValidationReport,
    extraction: Path,
    source_lines: Sequence[int],
) -> dict[str, str]:
    occurrences: dict[str, list[int]] = defaultdict(list)
    for source_id, line in zip(source_ids, source_lines):
        occurrences[source_id].append(line)
    for source_id, lines in occurrences.items():
        if source_id and len(lines) > 1:
            report.error(
                "ID_LANCAMENTO_DUPLICADO",
                file=extraction,
                line=", ".join(str(line) for line in lines[:20]),
                source_id=source_id,
                detail="O Id deve identificar um unico lancamento.",
            )

    valid_ids = [source_id for source_id in source_ids if source_id]
    preserve = bool(valid_ids) and all(source_id.isdigit() and len(source_id) <= 6 for source_id in valid_ids)
    result: dict[str, str] = {}
    if preserve:
        for source_id in valid_ids:
            result[source_id] = source_id
        return result

    next_code = 100000
    for source_id in valid_ids:
        if source_id in result:
            continue
        if next_code > 999999:
            raise ValueError("Quantidade de lancamentos ultrapassou o limite de codigos de 6 digitos.")
        result[source_id] = str(next_code)
        next_code += 1
    return result


def process(
    extraction_path: Path,
    account_relation_path: Path,
    model_path: Path,
    supplier_model_path: Path,
    supplier_primary_path: Path | None,
    supplier_complement_path: Path,
    *,
    default_account_origin: str,
    import_date: str,
    report: ValidationReport,
) -> tuple[ProcessResult, Table, Table, list[list[str]], int]:
    model = read_csv_table(model_path)
    if len(model.headers) != 14:
        raise ValueError(
            f"Modelo de Contas a Pagar possui {len(model.headers)} colunas; esperado: 14."
        )
    model_indexes = map_model_fields(model.headers, MODEL_FIELDS)

    supplier_template = read_csv_table(supplier_model_path)
    if len(supplier_template.headers) != 14:
        raise ValueError(
            "Modelo de Fornecedor deve possuir 14 colunas; "
            f"encontrado: {len(supplier_template.headers)}."
        )
    if supplier_complement_path.is_file():
        supplier_model_or_target = read_csv_table(supplier_complement_path)
        if [clean_text(value) for value in supplier_model_or_target.headers] != [
            clean_text(value) for value in supplier_template.headers
        ]:
            raise ValueError(
                "O cabecalho do arquivo complementar de fornecedores diverge do modelo."
            )
    else:
        supplier_model_or_target = Table(
            path=supplier_complement_path,
            sheet=supplier_complement_path.name,
            header_row=1,
            headers=list(supplier_template.headers),
            rows=[],
            delimiter=supplier_template.delimiter,
            encoding=supplier_template.encoding,
        )
    supplier_indexes = map_model_fields(supplier_model_or_target.headers, SUPPLIER_MODEL_FIELDS)
    supplier_observations_updated = update_existing_supplier_observations(
        supplier_model_or_target,
        supplier_indexes,
        import_date,
    )

    extraction, _ = locate_xlsx_table(
        extraction_path,
        EXTRACTION_SIGNATURE,
        minimum_score=7,
    )
    columns = source_columns(extraction.headers)
    account_relations = load_account_relations(account_relation_path, report)

    # Regra fixa definida para Contas a Pagar. O valor nao e extraido das
    # linhas do modelo nem relacionado pela aba FP do DE-PARA.
    payment_code = clean_text(FIXED_PAYMENT_CODE)
    account_origin = plan_account_code(default_account_origin)
    if not payment_code:
        raise ValueError("O codigo fixo da Forma de Pagamento esta vazio.")
    if not account_origin:
        raise ValueError(
            "A Conta de Origem padrao esta vazia ou invalida. "
            "Informe um codigo valido com --conta-origem-padrao."
        )
    print(
        f"Forma de pagamento fixa: {payment_code!r} "
        f"| conta de origem: {account_origin!r}"
    )

    supplier_paths: list[Path] = []
    if supplier_primary_path is not None and supplier_primary_path.is_file():
        supplier_paths.append(supplier_primary_path)
    if supplier_complement_path.is_file():
        supplier_paths.append(supplier_complement_path)
    suppliers = load_supplier_catalog(supplier_paths)

    prepared_rows: list[tuple[int, dict[str, Any]]] = []
    source_ids: list[str] = []
    source_lines: list[int] = []
    for line, raw_row in extraction.rows:
        values = {
            field_name: row_value(raw_row, index)
            for field_name, index in columns.items()
        }
        source_id = normalize_id(values["id"])
        values["_source_id"] = source_id
        prepared_rows.append((line, values))
        source_ids.append(source_id)
        source_lines.append(line)

    transaction_codes = assign_transaction_codes(source_ids, report, extraction_path, source_lines)
    raw_id_set = {source_id for source_id in source_ids if source_id}
    extra_relation_ids = sorted(set(account_relations) - raw_id_set)
    if extra_relation_ids:
        report.warning(
            "RELACOES_CONTA_SEM_LANCAMENTO",
            file=account_relation_path,
            origin=len(extra_relation_ids),
            detail=(
                "Existem Ids no arquivo de plano de contas que nao aparecem na extracao atual. "
                "Primeiros: " + ", ".join(extra_relation_ids[:20])
            ),
        )

    generated_by_key: dict[str, GeneratedSupplier] = {}
    output_rows: list[list[str]] = []
    source_total = Decimal("0")
    output_total = Decimal("0")
    confirmed_count = 0
    unconfirmed_count = 0

    required_indexes = required_output_indexes(model.headers)

    for line, values in prepared_rows:
        source_id = values["_source_id"]
        supplier_name = clean_text(values["supplier"])
        row_has_error = False

        if not source_id:
            report.error(
                "LANCAMENTO_SEM_ID",
                file=extraction_path,
                line=line,
                supplier=supplier_name,
                field=extraction.headers[columns["id"] or 0],
                detail="Linha preenchida sem Id.",
            )
            continue
        output_code = transaction_codes.get(source_id, "")
        if not output_code:
            report.error(
                "CODIGO_IMPORTACAO_NAO_GERADO",
                file=extraction_path,
                line=line,
                source_id=source_id,
                supplier=supplier_name,
            )
            continue

        account = account_relations.get(source_id)
        if account is None:
            report.error(
                "PLANO_CONTA_NAO_RELACIONADO",
                file=extraction_path,
                line=line,
                source_id=source_id,
                supplier=supplier_name,
                field="Codigo da Conta Destino",
                detail="Id nao localizado no arquivo de relacao do plano de contas.",
            )
            row_has_error = True

        if not supplier_name:
            report.error(
                "FORNECEDOR_SEM_NOME",
                file=extraction_path,
                line=line,
                source_id=source_id,
                field=extraction.headers[columns["supplier"] or 0],
                detail="Nao e possivel relacionar ou criar fornecedor sem nome.",
            )
            row_has_error = True
            supplier_code = ""
        else:
            supplier_code, match_method = match_supplier(suppliers, supplier_name)
            if not supplier_code and match_method.startswith("AMBIGUO"):
                report.error(
                    "FORNECEDOR_AMBIGUO",
                    file=extraction_path,
                    line=line,
                    source_id=source_id,
                    supplier=supplier_name,
                    detail=f"Mais de um codigo possivel no cadastro tratado ({match_method}).",
                )
                row_has_error = True
            elif not supplier_code:
                identity = compact(supplier_name)  # FORNECEDOR_IDENTIDADE_COMPACTA_V2026_08_25_1
                generated = generated_by_key.get(identity)
                if generated is None:
                    generated = GeneratedSupplier(
                        code=next_supplier_code(suppliers.used_codes),
                        name=supplier_name,
                    )
                    generated_by_key[identity] = generated
                    add_supplier_to_catalog(suppliers, generated.code, generated.name)
                supplier_code = generated.code

        confirmation = status_confirmation(values["status"])
        if not confirmation:
            report.error(
                "STATUS_NAO_RECONHECIDO",
                file=extraction_path,
                line=line,
                source_id=source_id,
                supplier=supplier_name,
                field=extraction.headers[columns["status"] or 0],
                origin=values["status"],
                detail="Esperado status pago/paga ou nao pago/nao paga/pendente.",
            )
            row_has_error = True

        due_date = date_out(values["due"], date1904=extraction.date1904)
        if not due_date:
            report.error(
                "DATA_VENCIMENTO_INVALIDA",
                file=extraction_path,
                line=line,
                source_id=source_id,
                supplier=supplier_name,
                field=extraction.headers[columns["due"] or 0],
                origin=values["due"],
                detail="A data de vencimento alimenta Lancamento, Competencia e Vencimento.",
            )
            row_has_error = True

        raw_paid_date = date_out(values["paid"], date1904=extraction.date1904)
        confirmation_date = raw_paid_date if confirmation == "S" else ""
        if confirmation == "S" and not confirmation_date:
            report.error(
                "CONTA_PAGA_SEM_DATA_PAGAMENTO",
                file=extraction_path,
                line=line,
                source_id=source_id,
                supplier=supplier_name,
                field="Data Pagamento",
                origin=values["paid"],
                detail="Contas confirmadas exigem Data Confirmacao.",
            )
            row_has_error = True

        amount_decimal = decimal_value(values["amount"])
        amount_text = money(values["amount"])
        if amount_decimal is None or not amount_text:
            report.error(
                "VALOR_TOTAL_INVALIDO",
                file=extraction_path,
                line=line,
                source_id=source_id,
                supplier=supplier_name,
                field=extraction.headers[columns["amount"] or 0],
                origin=values["amount"],
            )
            row_has_error = True
        else:
            source_total += amount_decimal.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        # Forma de pagamento e conta de origem sao padroes desta importacao.
        # Uma eventual coluna de forma na extracao e intencionalmente ignorada.

        if row_has_error or account is None or amount_decimal is None:
            continue

        output_row = ["" for _ in model.headers]
        values_to_set: dict[str, Any] = {
            "code": output_code,
            "launch": due_date,
            "competence": due_date,
            "due": due_date,
            "confirmation": confirmation_date,
            "supplier": supplier_code,
            "account_origin": account_origin,
            "account_destination": account.code,
            "payment": payment_code,
            "amount": amount_text,
            "interest": "",
            "paid_amount": amount_text if confirmation == "S" else "",
            "confirmed": confirmation,
            "observation": build_observation(
                values,
                account,
                import_date=import_date,
                original_id=source_id,
                output_code=output_code,
                unconfirmed_payment_date=raw_paid_date if confirmation == "N" else "",
            ),
        }
        for field_name, value in values_to_set.items():
            index = model_indexes.get(field_name)
            if index is not None:
                output_row[index] = value

        missing = [
            clean_text(model.headers[index])
            for index in required_indexes
            if not clean_text(row_value(output_row, index))
        ]
        if confirmation == "S":
            for field_name in ("due", "confirmation", "paid_amount"):
                index = model_indexes.get(field_name)
                if index is not None and not clean_text(row_value(output_row, index)):
                    missing.append(clean_text(model.headers[index]))
        if missing:
            report.error(
                "CAMPOS_OBRIGATORIOS_NAO_PREENCHIDOS",
                file=extraction_path,
                line=line,
                source_id=source_id,
                supplier=supplier_name,
                detail=", ".join(dict.fromkeys(missing)),
            )
            continue

        if len(output_code) > 6 or not output_code.isdigit():
            report.error(
                "CODIGO_LANCAMENTO_INVALIDO",
                file=extraction_path,
                line=line,
                source_id=source_id,
                origin=output_code,
                detail="Codigo de importacao deve ser numerico e possuir no maximo 6 digitos.",
            )
            continue

        output_rows.append([output_text(value, model.delimiter) for value in output_row])
        output_total += amount_decimal.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        if confirmation == "S":
            confirmed_count += 1
        else:
            unconfirmed_count += 1

    generated_suppliers = list(generated_by_key.values())
    supplier_rows = [
        build_supplier_row(
            supplier_model_or_target.headers,
            supplier_indexes,
            supplier,
            import_date,
        )
        for supplier in generated_suppliers
    ]
    for supplier, row in zip(generated_suppliers, supplier_rows):
        missing = validate_supplier_row(supplier_model_or_target.headers, supplier_indexes, row)
        if missing:
            report.error(
                "FORNECEDOR_COMPLEMENTAR_INCOMPLETO",
                file=supplier_model_path,
                source_id=supplier.code,
                supplier=supplier.name,
                detail=", ".join(missing),
            )

    if len(output_rows) != len(prepared_rows) and report.error_count == 0:
        report.error(
            "CONTAGEM_FINAL_DIVERGENTE",
            file=extraction_path,
            origin=len(prepared_rows),
            compared=len(output_rows),
            detail="Nem todas as linhas preenchidas da extracao chegaram ao CSV final.",
        )
    if source_total != output_total and report.error_count == 0:
        report.error(
            "TOTAL_FINAL_DIVERGENTE",
            file=extraction_path,
            origin=money(source_total),
            compared=money(output_total),
            detail="A soma de Valor Titulo divergiu da soma da extracao.",
        )

    return (
        ProcessResult(
            output_rows=output_rows,
            generated_suppliers=generated_suppliers,
            source_count=len(prepared_rows),
            confirmed_count=confirmed_count,
            unconfirmed_count=unconfirmed_count,
            source_total=source_total,
            output_total=output_total,
        ),
        model,
        supplier_model_or_target,
        supplier_rows,
        supplier_observations_updated,
    )


# ---------------------------------------------------------------------------
# Escrita atomica, validacao final e relatorio
# ---------------------------------------------------------------------------

REPLACE_ATTEMPTS = 6
REPLACE_DELAY_SECONDS = 0.5
COPY_CHUNK_SIZE = 1024 * 1024


class OutputFileInUseError(PermissionError):
    """Arquivo final nao pode ser substituido porque esta aberto ou protegido."""

    def __init__(self, path: Path, original: BaseException | None = None) -> None:
        self.path = path
        self.original = original
        super().__init__(
            f"Nao foi possivel atualizar {path}. O arquivo provavelmente esta "
            "aberto no Excel, no visualizador do Explorador de Arquivos ou em "
            "outro programa. Feche o arquivo e execute o script novamente. "
            "A versao anterior foi preservada."
        )


def is_file_lock_error(exc: BaseException) -> bool:
    """Reconhece erros de compartilhamento/permissao comuns no Windows."""
    winerror = getattr(exc, "winerror", None)
    error_number = getattr(exc, "errno", None)
    return isinstance(exc, PermissionError) or winerror in {5, 32, 33} or error_number in {
        errno.EACCES,
        errno.EPERM,
        errno.EBUSY,
    }


def unlink_with_retry(path: Path, *, ignore_missing: bool = True) -> None:
    """Remove arquivo temporario, tolerando travas transitorias."""
    last_error: OSError | None = None
    for attempt in range(REPLACE_ATTEMPTS):
        try:
            path.unlink(missing_ok=ignore_missing)
            return
        except FileNotFoundError:
            if ignore_missing:
                return
            raise
        except OSError as exc:
            if not is_file_lock_error(exc):
                raise
            last_error = exc
            if attempt + 1 < REPLACE_ATTEMPTS:
                time.sleep(REPLACE_DELAY_SECONDS)
    if last_error is not None:
        raise OutputFileInUseError(path, last_error) from last_error


def files_are_identical(first: Path, second: Path) -> bool:
    """Compara dois arquivos sem carrega-los integralmente na memoria."""
    try:
        if not first.is_file() or not second.is_file():
            return False
        if first.stat().st_size != second.stat().st_size:
            return False
        with first.open("rb") as left, second.open("rb") as right:
            while True:
                left_chunk = left.read(COPY_CHUNK_SIZE)
                right_chunk = right.read(COPY_CHUNK_SIZE)
                if left_chunk != right_chunk:
                    return False
                if not left_chunk:
                    return True
    except OSError:
        return False


def ensure_target_replaceable(path: Path) -> None:
    """Falha cedo e com mensagem clara quando o arquivo final esta em uso."""
    if not path.exists():
        return
    last_error: OSError | None = None
    for attempt in range(REPLACE_ATTEMPTS):
        descriptor: int | None = None
        try:
            descriptor = os.open(path, os.O_RDWR)
            return
        except OSError as exc:
            if not is_file_lock_error(exc):
                raise
            last_error = exc
            if attempt + 1 < REPLACE_ATTEMPTS:
                time.sleep(REPLACE_DELAY_SECONDS)
        finally:
            if descriptor is not None:
                os.close(descriptor)
    if last_error is not None:
        raise OutputFileInUseError(path, last_error) from last_error


def replace_with_retry(source: Path, target: Path) -> None:
    """Substitui um arquivo com tentativas para antivirus/sincronizacao transitoria."""
    last_error: OSError | None = None
    for attempt in range(REPLACE_ATTEMPTS):
        try:
            os.replace(source, target)
            return
        except OSError as exc:
            if not is_file_lock_error(exc):
                raise
            last_error = exc
            if attempt + 1 < REPLACE_ATTEMPTS:
                time.sleep(REPLACE_DELAY_SECONDS)
    if last_error is not None:
        raise OutputFileInUseError(target, last_error) from last_error


def cleanup_staged_files(paths: Iterable[Path | None]) -> None:
    """Remove arquivos .novo/.anterior que nao devem permanecer na pasta saida."""
    for path in paths:
        if path is None or not path.exists():
            continue
        try:
            unlink_with_retry(path)
        except OSError:
            # Nao mascara o erro principal por causa de uma limpeza secundaria.
            pass


def commit_staged_files(
    staged_and_final: Sequence[tuple[Path | None, Path]],
) -> tuple[list[Path], list[Path]]:
    """Publica varios CSVs com comparacao, preflight, backup e rollback."""
    pending: list[tuple[Path, Path]] = []
    unchanged: list[Path] = []

    for staged, final in staged_and_final:
        if staged is None:
            continue
        if final.exists() and files_are_identical(staged, final):
            unlink_with_retry(staged)
            unchanged.append(final)
        else:
            pending.append((staged, final))

    if not pending:
        return [], unchanged

    backups: dict[Path, Path | None] = {}
    committed: list[Path] = []
    try:
        # Verifica todos antes de alterar o primeiro, evitando atualizacao parcial.
        for _, final in pending:
            ensure_target_replaceable(final)

        for _, final in pending:
            backup = final.with_name(final.name + ".anterior")
            unlink_with_retry(backup)
            if final.exists():
                shutil.copy2(final, backup)
                backups[final] = backup
            else:
                backups[final] = None

        for staged, final in pending:
            replace_with_retry(staged, final)
            committed.append(final)

        return committed, unchanged
    except Exception as commit_error:
        rollback_errors: list[str] = []
        for final in reversed(committed):
            backup = backups.get(final)
            try:
                if backup is not None and backup.exists():
                    replace_with_retry(backup, final)
                elif final.exists():
                    unlink_with_retry(final)
            except Exception as rollback_error:
                rollback_errors.append(f"{final}: {rollback_error}")
        if rollback_errors:
            raise RuntimeError(
                "Falha ao publicar os CSVs e tambem ao restaurar arquivo(s): "
                + " | ".join(rollback_errors)
            ) from commit_error
        raise
    finally:
        cleanup_staged_files(staged for staged, _ in pending)
        cleanup_staged_files(backup for backup in backups.values() if backup is not None)

def write_csv_atomic(
    path: Path,
    headers: Sequence[Any],
    rows: Iterable[Sequence[Any]],
    *,
    delimiter: str,
    encoding: str,
) -> None:
    if not path.parent.is_dir():
        raise FileNotFoundError(f"Pasta de destino nao existe: {path.parent}")
    temporary = path.with_name(path.name + ".tmp")
    try:
        with temporary.open("w", encoding=encoding, newline="") as handle:
            writer = csv.writer(handle, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
            writer.writerow([encoding_safe(value, encoding, delimiter) for value in headers])
            for row in rows:
                writer.writerow([encoding_safe(value, encoding, delimiter) for value in row])
        if path.exists() and files_are_identical(temporary, path):
            unlink_with_retry(temporary)
        else:
            ensure_target_replaceable(path)
            replace_with_retry(temporary, path)
    finally:
        cleanup_staged_files([temporary])


def stage_csv(
    final_path: Path,
    headers: Sequence[Any],
    rows: Iterable[Sequence[Any]],
    *,
    delimiter: str,
    encoding: str,
) -> Path:
    if not final_path.parent.is_dir():
        raise FileNotFoundError(f"Pasta de destino nao existe: {final_path.parent}")
    staged = final_path.with_name(final_path.name + ".novo")
    if staged.exists():
        unlink_with_retry(staged)
    try:
        with staged.open("w", encoding=encoding, newline="") as handle:
            writer = csv.writer(handle, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
            writer.writerow([encoding_safe(value, encoding, delimiter) for value in headers])
            for row in rows:
                writer.writerow([encoding_safe(value, encoding, delimiter) for value in row])
        return staged
    except Exception:
        cleanup_staged_files([staged])
        raise


def validate_csv_output(
    path: Path,
    expected_headers: Sequence[Any],
    expected_rows: int,
    *,
    expected_observation_suffix: str | None = None,
) -> None:
    table = read_csv_table(path)
    normalized_expected = [clean_text(value) for value in expected_headers]
    normalized_actual = [clean_text(value) for value in table.headers]
    if normalized_actual != normalized_expected:
        raise ValueError(f"Cabecalho final divergiu do modelo em {path.name}")
    if len(table.rows) != expected_rows:
        raise ValueError(
            f"Quantidade final incorreta em {path.name}: {len(table.rows)}; esperado: {expected_rows}."
        )
    observation_index = find_header(table.headers, ["Observacao"], required=False)
    for line, row in table.rows:
        if len(row) != len(expected_headers):
            raise ValueError(
                f"Linha {line} de {path.name} possui {len(row)} colunas; "
                f"esperado: {len(expected_headers)}."
            )
        for value in row:
            if any(character in value for character in ('"', "'", "\\", "\r", "\n", "\t")):
                raise ValueError(f"Linha {line} de {path.name} contem caractere proibido.")
        if expected_observation_suffix and observation_index is not None:
            observation = clean_text(row_value(row, observation_index))
            if not observation.endswith(expected_observation_suffix):
                raise ValueError(
                    f"Linha {line} de {path.name}: Observacao nao termina com "
                    f"{expected_observation_suffix}."
                )


def merge_supplier_rows(
    target_table: Table,
    new_rows: Sequence[Sequence[Any]],
) -> list[list[Any]]:
    existing = [list(row) for _, row in target_table.rows]
    width = len(target_table.headers)
    result: list[list[Any]] = []
    for row in [*existing, *new_rows]:
        adjusted = list(row[:width]) + [""] * max(0, width - len(row))
        result.append(adjusted[:width])
    return result


def write_validation(path: Path, report: ValidationReport) -> None:
    findings = [issue for issue in report.issues if issue.level in {"ERRO", "AVISO"}]
    if not findings:
        if path.exists():
            unlink_with_retry(path)
        return
    if not path.parent.is_dir():
        raise FileNotFoundError(f"Pasta de validacao nao existe: {path.parent}")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Validacao"
    headers = [
        "Nivel",
        "Tipo",
        "Arquivo",
        "Linha origem",
        "Id origem",
        "Favorecido",
        "Campo",
        "Valor origem",
        "Valor comparado",
        "Detalhe",
    ]
    worksheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for issue in findings:
        worksheet.append(
            [
                issue.level,
                issue.category,
                issue.file,
                issue.line,
                issue.source_id,
                issue.supplier,
                issue.field,
                issue.origin,
                issue.compared,
                issue.detail,
            ]
        )
        row_number = worksheet.max_row
        color = "F4CCCC" if issue.level == "ERRO" else "FCE5CD"
        worksheet.cell(row=row_number, column=1).fill = PatternFill("solid", fgColor=color)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    widths = {
        "A": 10,
        "B": 38,
        "C": 34,
        "D": 14,
        "E": 16,
        "F": 36,
        "G": 30,
        "H": 35,
        "I": 35,
        "J": 90,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    temporary = path.with_name(path.name + ".tmp.xlsx")
    try:
        workbook.save(temporary)
        workbook.close()
        if path.exists() and files_are_identical(temporary, path):
            unlink_with_retry(temporary)
        else:
            ensure_target_replaceable(path)
            replace_with_retry(temporary, path)
    finally:
        cleanup_staged_files([temporary])


# ---------------------------------------------------------------------------
# Linha de comando
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Tratamento automatico de Contas a Pagar da Laser Rosa"
    )
    parser.add_argument("--entrada", type=Path, help="Extracao Contas a Pagar.xlsx")
    parser.add_argument(
        "--plano-contas",
        type=Path,
        help="Arquivo de relacao Id -> codigo do plano de contas",
    )
    parser.add_argument("--modelo", type=Path, help="modeloImportacaoContasPagar.csv")
    parser.add_argument(
        "--fornecedores",
        type=Path,
        help="planilhaTratadaFornecedor.csv/.xlsx existente, opcional",
    )
    parser.add_argument(
        "--modelo-fornecedores",
        type=Path,
        help="modeloImportacaoFornecedor.csv",
    )
    parser.add_argument(
        "--fornecedores-complementares",
        type=Path,
        help="CSV complementar de fornecedores ausentes",
    )
    # Argumentos antigos mantidos apenas para compatibilidade com comandos salvos.
    # A Forma de Pagamento agora e fixa e nao depende do DE-PARA.
    parser.add_argument(
        "--de-para",
        dest="depara",
        type=Path,
        help="Opcao obsoleta; o DE-PARA nao e usado para a Forma de Pagamento",
    )
    parser.add_argument(
        "--forma-pagamento-padrao",
        default="",
        help=(
            "Opcao obsoleta; o valor informado e ignorado. "
            f"O codigo fixo e {FIXED_PAYMENT_CODE!r}."
        ),
    )
    parser.add_argument(
        "--listar-formas-pagamento",
        action="store_true",
        help="Mostra o codigo fixo da Forma de Pagamento e encerra",
    )
    parser.add_argument(
        "--nao-interativo",
        action="store_true",
        help="Opcao obsoleta mantida por compatibilidade; o script nao solicita escolha",
    )
    parser.add_argument(
        "--conta-origem-padrao",
        default="1.1.1.001",
        help="Codigo fixo da Conta de Origem; padrao: 1.1.1.001",
    )
    parser.add_argument(
        "--data-importacao",
        help="Data adicionada ao fim da Observacao, em DD/MM/AAAA; padrao: hoje",
    )
    parser.add_argument("--saida", type=Path, help="CSV final de Contas a Pagar")
    parser.add_argument("--validacao", type=Path, help="XLSX de erros/avisos")
    return parser


def main() -> int:
    print(f"Laser Rosa - Contas a Pagar | versao {VERSION}")
    print(f"Arquivo executado: {Path(__file__).resolve()}")
    print(f"Identificador do arquivo: {script_hash()}")
    args = build_parser().parse_args()
    report = ValidationReport()
    validation_path: Path | None = None
    staged_supplier: Path | None = None
    staged_output: Path | None = None
    try:
        configure_project_layout()

        if args.listar_formas_pagamento:
            account_origin = plan_account_code(args.conta_origem_padrao)
            print("Forma de pagamento fixa para Contas a Pagar:")
            print(
                f"  - codigo={FIXED_PAYMENT_CODE!r} "
                f"| conta_origem={account_origin!r}"
            )
            return 0

        obsolete_payment = clean_text(args.forma_pagamento_padrao)
        if obsolete_payment and norm(obsolete_payment) != norm(FIXED_PAYMENT_CODE):
            print(
                "AVISO: --forma-pagamento-padrao foi ignorado. "
                f"O valor fixo desta importacao e {FIXED_PAYMENT_CODE!r}."
            )

        explicit_extraction = resolve_argument(args.entrada, INPUT_DIR)
        explicit_accounts = resolve_argument(args.plano_contas, INPUT_DIR)
        extraction = discover_extraction(INPUT_DIR, explicit_extraction)
        account_relation = discover_account_relation(INPUT_DIR, explicit_accounts)
        model = resolve_argument(args.modelo, INPUT_DIR) or resolve_variant(
            "modeloImportacaoContasPagar", {".csv"}, INPUT_DIR
        )
        supplier_model = resolve_argument(args.modelo_fornecedores, INPUT_DIR) or resolve_variant(
            "modeloImportacaoFornecedor", {".csv"}, INPUT_DIR
        )
        supplier_primary = resolve_argument(args.fornecedores, OUTPUT_DIR) or resolve_supplier_primary(OUTPUT_DIR)
        supplier_complement = resolve_argument(args.fornecedores_complementares, OUTPUT_DIR) or (
            OUTPUT_DIR / "planilhaTratadaFornecedorContasPagar.csv"
        )
        output = resolve_argument(args.saida, OUTPUT_DIR) or (
            OUTPUT_DIR / "planilhaTratadaContasPagar.csv"
        )
        validation_path = resolve_argument(args.validacao, OUTPUT_DIR) or (
            OUTPUT_DIR / "validacaoContasPagar.xlsx"
        )

        assert model is not None
        assert supplier_model is not None

        # Remove restos de uma execucao anterior interrompida. O CSV final nao e
        # tocado nesta etapa.
        cleanup_staged_files(
            [
                output.with_name(output.name + ".novo"),
                supplier_complement.with_name(supplier_complement.name + ".novo"),
                output.with_name(output.name + ".anterior"),
                supplier_complement.with_name(supplier_complement.name + ".anterior"),
            ]
        )

        if args.data_importacao:
            parsed_import_date = parse_date(args.data_importacao)
            if parsed_import_date is None:
                raise ValueError("--data-importacao deve estar no formato DD/MM/AAAA.")
            import_date = parsed_import_date.strftime("%d/%m/%Y")
        else:
            import_date = date.today().strftime("%d/%m/%Y")

        print(f"Data da importacao: {import_date}")

        result, model_table, supplier_target_table, supplier_rows, supplier_observations_updated = process(
            extraction,
            account_relation,
            model,
            supplier_model,
            supplier_primary,
            supplier_complement,
            default_account_origin=clean_text(args.conta_origem_padrao),
            import_date=import_date,
            report=report,
        )

        if report.error_count:
            write_validation(validation_path, report)
            print(
                f"BLOQUEADO [v{VERSION}]: {report.error_count} erro(s) e "
                f"{report.warning_count} aviso(s).",
                file=sys.stderr,
            )
            print(
                "Nenhum CSV final foi criado ou sobrescrito nesta execucao.",
                file=sys.stderr,
            )
            print(f"Validacao: {validation_path}", file=sys.stderr)
            return 2

        if result.generated_suppliers or supplier_observations_updated:
            merged_supplier_rows = merge_supplier_rows(supplier_target_table, supplier_rows)
            staged_supplier = stage_csv(
                supplier_complement,
                supplier_target_table.headers,
                merged_supplier_rows,
                delimiter=supplier_target_table.delimiter,
                encoding=supplier_target_table.encoding,
            )
            validate_csv_output(
                staged_supplier,
                supplier_target_table.headers,
                len(merged_supplier_rows),
                expected_observation_suffix=f"Importação {import_date}",
            )

        staged_output = stage_csv(
            output,
            model_table.headers,
            result.output_rows,
            delimiter=model_table.delimiter,
            encoding=model_table.encoding,
        )
        validate_csv_output(
            staged_output,
            model_table.headers,
            len(result.output_rows),
            expected_observation_suffix=f"Importação {import_date}",
        )

        # Limpa/grava a validacao antes de publicar os CSVs. Assim, um arquivo
        # de validacao aberto no Excel nao causa uma mensagem incorreta depois
        # que os CSVs ja tiverem sido atualizados.
        write_validation(validation_path, report)

        updated_files, unchanged_files = commit_staged_files(
            [
                (staged_output, output),
                (staged_supplier, supplier_complement),
            ]
        )
        staged_output = None
        staged_supplier = None

        print(
            f"OK [v{VERSION}]: {len(result.output_rows)} conta(s) exportada(s) -> {output}"
        )
        print(
            f"Confirmadas: {result.confirmed_count} | Nao confirmadas: "
            f"{result.unconfirmed_count} | Total: {money(result.output_total)}"
        )
        if result.generated_suppliers:
            print(
                f"Fornecedores complementares: {len(result.generated_suppliers)} -> "
                f"{supplier_complement}"
            )
            print(
                "Ordem de importacao: fornecedores complementares primeiro; depois Contas a Pagar."
            )
        if supplier_observations_updated:
            print(
                "Observacoes de fornecedores atualizadas para "
                f"{import_date}: {supplier_observations_updated} -> "
                f"{supplier_complement}"
            )
        if output in unchanged_files:
            print("Contas a Pagar: o CSV existente ja estava atualizado; sobrescrita dispensada.")
        if supplier_complement in unchanged_files and (
            result.generated_suppliers or supplier_observations_updated
        ):
            print("Fornecedores complementares: o CSV existente ja estava atualizado.")
        if report.warning_count:
            print(
                f"Validacao: {report.warning_count} aviso(s) -> {validation_path}"
            )
        else:
            print("Validacao: nenhum erro ou aviso; arquivo de validacao nao foi criado.")
        return 0

    except OutputFileInUseError as exc:
        cleanup_staged_files([staged_output, staged_supplier])
        print(f"ERRO [v{VERSION}]: ARQUIVO DE SAIDA EM USO", file=sys.stderr)
        print(str(exc), file=sys.stderr)
        print(
            "Nenhum CSV final foi criado ou sobrescrito nesta execucao.",
            file=sys.stderr,
        )
        print(
            "Validacao: nao criada, pois a falha e de acesso ao arquivo e nao "
            "de consistencia dos dados.",
            file=sys.stderr,
        )
        return 3

    except Exception as exc:
        cleanup_staged_files([staged_output, staged_supplier])
        report.error("FALHA_FATAL", detail=f"{type(exc).__name__}: {exc}")
        if validation_path is not None:
            try:
                write_validation(validation_path, report)
            except Exception as validation_exc:
                print(
                    f"ERRO ao gravar validacao: {validation_exc}",
                    file=sys.stderr,
                )
        print(f"ERRO [v{VERSION}]: {type(exc).__name__}: {exc}", file=sys.stderr)
        if validation_path is not None:
            print(f"Validacao: {validation_path}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
