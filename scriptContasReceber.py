#!/usr/bin/env python3
"""Tratamento de Contas a Receber - Laser Rosa.

Estrutura esperada:
    raiz/
      scriptContasReceber.py
      entrada/
        <extracao de contas a receber>.xlsx
        modeloImportacaoContasReceber.csv
        modeloImportacaoCliente.csv              # usado se houver cliente ausente
        DE-PARA.xlsx                             # aba FP/Formas de Pagamento
      saida/
        planilhaTratadaCliente.csv               # fonte do codigo de cliente
        planilhaTratadaContasReceber.csv         # gerado por este script
        planilhaTratadaClienteVendaPlano.csv     # fonte complementar e destino de novos clientes
        validacaoContasReceber.xlsx              # somente se houver pendencias

Principios do projeto:
- modelos servem somente de layout/referencia; suas linhas nunca sao fonte de dados;
- codigos de cliente vem de planilhaTratadaCliente e, como segunda fonte, da
  planilhaTratadaClienteVendaPlano; clientes ainda ausentes sao incluidos nesta
  segunda planilha com base no cabecalho do modeloImportacaoCliente;
- o arquivo unico DE-PARA e usado na aba FP/Formas de Pagamento;
- arquivos de validacao so existem quando ha erro/aviso de revisao;
- 'Importacao DD/MM/AAAA' fica sempre no fim da observacao;
- remove aspas simples/duplas, barra invertida, controles e quebras de linha;
- nao cria automaticamente as pastas entrada/saida;
- detecta a extracao pelo conteudo/cabecalhos, mesmo com nome variado e styles.xml corrompido.

Uso normal:
    py scriptContasReceber.py
"""

from __future__ import annotations

import argparse
import csv
import posixpath
import re
import sys
import unicodedata
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path, PurePosixPath
from typing import Any, Iterable, Iterator, Sequence
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook

VERSION = "2026-08-26.8"
FORBIDDEN = str.maketrans({'"': '', "'": '', "\\": ''})
NULL_WORDS = {"", "null", "none", "nan", "nat", "\\n"}
CELL_REF_RE = re.compile(r"([A-Z]+)(\d+)")

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR
INPUT_DIR = PROJECT_ROOT / "entrada"
OUTPUT_DIR = PROJECT_ROOT / "saida"


# ---------------------------------------------------------------------------
# Limpeza / normalizacao
# ---------------------------------------------------------------------------
def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    if text.strip().casefold() in NULL_WORDS:
        return ""
    text = text.translate(FORBIDDEN)
    text = "".join(ch for ch in text if unicodedata.category(ch) not in {"Cc", "Cf"})
    text = text.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    return " ".join(text.split()).strip()


def ascii_fold(value: Any) -> str:
    text = unicodedata.normalize("NFKD", clean_text(value))
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def norm(value: Any) -> str:
    text = ascii_fold(value).casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def digits(value: Any) -> str:
    return "".join(ch for ch in clean_text(value) if ch.isdigit())


def column_index(reference: str) -> int:
    match = CELL_REF_RE.match(reference)
    if not match:
        return -1
    value = 0
    for char in match.group(1):
        value = value * 26 + (ord(char) - 64)
    return value - 1


class XlsxReader:
    """Le XLSX diretamente do XML, ignorando styles.xml corrompido."""

    MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

    def __init__(self, path: Path):
        self.path = path
        self.archive = ZipFile(path)
        self.shared_strings = self._load_shared_strings()
        self.sheets = self._load_workbook_info()

    def close(self) -> None:
        self.archive.close()

    def __enter__(self) -> "XlsxReader":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    def _load_shared_strings(self) -> list[str]:
        if "xl/sharedStrings.xml" not in self.archive.namelist():
            return []
        root = ET.fromstring(self.archive.read("xl/sharedStrings.xml"))
        result: list[str] = []
        for item in root.findall(f"{{{self.MAIN_NS}}}si"):
            parts = [node.text or "" for node in item.iter(f"{{{self.MAIN_NS}}}t")]
            result.append("".join(parts))
        return result

    def _load_workbook_info(self) -> dict[str, str]:
        workbook = ET.fromstring(self.archive.read("xl/workbook.xml"))
        relations_root = ET.fromstring(self.archive.read("xl/_rels/workbook.xml.rels"))
        relations: dict[str, str] = {}
        for relation in relations_root.findall(f"{{{self.PKG_REL_NS}}}Relationship"):
            relation_id = relation.attrib.get("Id", "")
            target = relation.attrib.get("Target", "")
            if not relation_id or not target:
                continue
            if target.startswith("/"):
                resolved = target.lstrip("/")
            else:
                resolved = posixpath.normpath(str(PurePosixPath("xl") / target))
            relations[relation_id] = resolved

        sheets: dict[str, str] = {}
        sheets_node = workbook.find(f"{{{self.MAIN_NS}}}sheets")
        if sheets_node is not None:
            for sheet in sheets_node.findall(f"{{{self.MAIN_NS}}}sheet"):
                name = sheet.attrib.get("name", "")
                relation_id = sheet.attrib.get(f"{{{self.REL_NS}}}id", "")
                if name and relation_id in relations:
                    sheets[name] = relations[relation_id]
        return sheets

    def iter_rows(
        self,
        sheet_name: str,
        min_row: int = 1,
        max_col: int = 120,
        max_row: int | None = None,
    ) -> Iterator[tuple[int, list[Any]]]:
        if sheet_name in self.sheets:
            sheet_path = self.sheets[sheet_name]
        else:
            lowered = {name.casefold(): path for name, path in self.sheets.items()}
            if sheet_name.casefold() in lowered:
                sheet_path = lowered[sheet_name.casefold()]
            elif self.sheets:
                sheet_path = next(iter(self.sheets.values()))
            else:
                raise KeyError(f"Nenhuma aba encontrada em {self.path}")

        with self.archive.open(sheet_path) as handle:
            for _, element in ET.iterparse(handle, events=("end",)):
                if element.tag != f"{{{self.MAIN_NS}}}row":
                    continue
                try:
                    row_number = int(element.attrib.get("r", "0"))
                except ValueError:
                    row_number = 0
                if row_number < min_row:
                    element.clear()
                    continue
                if max_row is not None and row_number > max_row:
                    element.clear()
                    break

                values: list[Any] = [None] * max_col
                for cell in element.findall(f"{{{self.MAIN_NS}}}c"):
                    reference = cell.attrib.get("r", "")
                    index = column_index(reference)
                    if index < 0 or index >= max_col:
                        continue
                    cell_type = cell.attrib.get("t", "")
                    value_node = cell.find(f"{{{self.MAIN_NS}}}v")
                    inline_node = cell.find(f"{{{self.MAIN_NS}}}is")
                    value: Any = None
                    if cell_type == "inlineStr" and inline_node is not None:
                        value = "".join(node.text or "" for node in inline_node.iter(f"{{{self.MAIN_NS}}}t"))
                    elif value_node is not None and value_node.text is not None:
                        raw = value_node.text
                        if cell_type == "s":
                            try:
                                value = self.shared_strings[int(raw)]
                            except (ValueError, IndexError):
                                value = ""
                        elif cell_type == "b":
                            value = raw == "1"
                        else:
                            value = raw
                    values[index] = value
                yield row_number, values
                element.clear()


def normalize_name(value: Any) -> str:
    return " ".join(re.sub(r"[^A-Z0-9]+", " ", ascii_fold(value).upper()).split())


def normalize_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else format(value, ".15g")
    text = clean_text(value)
    if re.fullmatch(r"\d+[.,]0+", text):
        return re.split(r"[.,]", text, 1)[0]
    return text


def money(value: Any) -> str:
    """Retorna decimal brasileiro sem separador de milhar."""
    if value is None or clean_text(value) == "":
        return "0"
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        dec = Decimal(str(value))
    else:
        text = clean_text(value).replace("R$", "").replace(" ", "")
        if not text:
            return "0"
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        else:
            text = text.replace(",", ".")
        try:
            dec = Decimal(text)
        except InvalidOperation:
            return ""
    quantized = dec.quantize(Decimal("0.01"))
    result = format(quantized, "f").replace(".", ",")
    result = result.rstrip("0").rstrip(",") if "," in result else result
    return result or "0"


def parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def date_out(value: Any) -> str:
    parsed = parse_date(value)
    return parsed.strftime("%d/%m/%Y") if parsed else ""


# ---------------------------------------------------------------------------
# Estrutura do projeto / descoberta
# ---------------------------------------------------------------------------
def _layout_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", ascii_fold(value).casefold())


def _existing_child(root: Path, logical_name: str) -> Path | None:
    wanted = _layout_key(logical_name)
    if not root.is_dir():
        return None
    for child in root.iterdir():
        if child.is_dir() and _layout_key(child.name) == wanted:
            return child.resolve()
    return None


def configure_project_layout() -> tuple[Path, Path, Path]:
    global PROJECT_ROOT, INPUT_DIR, OUTPUT_DIR
    candidates: list[tuple[int, int, Path, Path, Path]] = []
    for root in (SCRIPT_DIR, *SCRIPT_DIR.parents):
        root = root.resolve()
        entrada = _existing_child(root, "entrada")
        saida = _existing_child(root, "saida")
        if not entrada or not saida:
            continue
        score = 0
        input_names = [_layout_key(p.stem) for p in entrada.iterdir() if p.is_file()]
        output_names = [_layout_key(p.stem) for p in saida.iterdir() if p.is_file()]
        if any(re.fullmatch(r"depara\d*", name) for name in input_names):
            score += 40
        if any(name.startswith("modeloimportacaocontasreceber") for name in input_names):
            score += 20
        if any(name.startswith("planilhatratadacliente") for name in output_names):
            score += 20
        score += min(len(input_names), 10)
        candidates.append((score, -len(root.parts), root, entrada, saida))
    if not candidates:
        raise FileNotFoundError(
            "Nao encontrei uma raiz existente contendo simultaneamente 'entrada' e 'saida' "
            "(ou 'saída') a partir da pasta do script e seus diretorios-pai."
        )
    _, _, PROJECT_ROOT, INPUT_DIR, OUTPUT_DIR = max(candidates, key=lambda x: (x[0], x[1]))
    return PROJECT_ROOT, INPUT_DIR, OUTPUT_DIR


def newest(paths: Sequence[Path]) -> Path:
    return max(paths, key=lambda p: p.stat().st_mtime)


def resolve_variant(stem: str, suffixes: set[str], directory: Path, *, required: bool = True) -> Path | None:
    exacts = [directory / f"{stem}{s}" for s in suffixes]
    for path in exacts:
        if path.is_file():
            return path.resolve()
    wanted = _layout_key(stem)
    candidates = [
        p for p in directory.iterdir()
        if p.is_file() and not p.name.startswith("~$") and p.suffix.lower() in suffixes
        and _layout_key(p.stem).startswith(wanted)
    ]
    if candidates:
        return newest(candidates).resolve()
    if required:
        raise FileNotFoundError(f"Arquivo {stem} nao encontrado em {directory}")
    return None


def resolve_depara(directory: Path) -> Path:
    candidates = [
        p for p in directory.iterdir()
        if p.is_file() and p.suffix.lower() in {".xlsx", ".xls"}
        and re.fullmatch(r"depara\d*", _layout_key(p.stem))
    ]
    if not candidates:
        raise FileNotFoundError(f"Arquivo unico DE-PARA nao encontrado em {directory}")
    exact = [p for p in candidates if _layout_key(p.stem) == "depara"]
    return newest(exact or candidates).resolve()


def read_csv(path: Path) -> tuple[list[str], list[list[str]], str, str]:
    raw = path.read_bytes()
    encoding = "utf-8-sig"
    text = None
    for enc in ("utf-8-sig", "cp1252", "latin1"):
        try:
            text = raw.decode(enc)
            encoding = enc
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError(f"Nao foi possivel ler {path}")
    first = next((line for line in text.splitlines() if line.strip()), "")
    delimiter = ";" if ";" in first else ","
    rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    if not rows:
        raise ValueError(f"Arquivo vazio: {path}")
    return rows[0], rows[1:], delimiter, encoding


def read_xlsx_headers(path: Path) -> tuple[str, list[Any], int, int]:
    """Localiza a linha de cabecalho real sem carregar styles.xml."""
    required = {"conta", "cliente", "cpf", "data de vencimento", "valor total", "status do pagamento"}
    best: tuple[int, str, int, list[Any]] | None = None
    with XlsxReader(path) as xlsx:
        for sheet_name in xlsx.sheets:
            for row_number, row in xlsx.iter_rows(sheet_name, min_row=1, max_col=120, max_row=50):
                normalized = {norm(value) for value in row if clean_text(value)}
                score = sum(any(req == header or req in header for header in normalized) for req in required)
                candidate = (score, sheet_name, row_number, row)
                if best is None or score > best[0]:
                    best = candidate
    if best is None or best[0] < 5:
        raise ValueError(f"Nao identifiquei a extracao de Contas a Receber em {path.name}")
    return best[1], best[3], best[0], best[2]


def discover_extraction(directory: Path, explicit: Path | None = None) -> Path:
    if explicit:
        path = explicit if explicit.is_absolute() else directory / explicit
        if not path.is_file():
            raise FileNotFoundError(path)
        # Tambem valida o layout quando o arquivo foi informado explicitamente.
        read_xlsx_headers(path)
        return path.resolve()

    matches: list[tuple[int, int, float, Path]] = []
    for path in directory.iterdir():
        if not path.is_file() or path.suffix.lower() != ".xlsx" or path.name.startswith("~$"):
            continue
        key = _layout_key(path.stem)
        if key.startswith("modeloimportacao") or re.fullmatch(r"depara\d*", key):
            continue
        try:
            _, _, score, _ = read_xlsx_headers(path)
        except Exception:
            continue
        name_hint = 1 if "contasareceber" in key or ("contas" in key and "receber" in key) else 0
        matches.append((score, name_hint, path.stat().st_mtime, path))

    if not matches:
        raise FileNotFoundError(f"Extracao de Contas a Receber nao encontrada em {directory}")
    return max(matches, key=lambda item: (item[0], item[1], item[2]))[3].resolve()


# ---------------------------------------------------------------------------
# Tabelas auxiliares
# ---------------------------------------------------------------------------
def find_header(headers: Sequence[Any], aliases: Sequence[str], required: bool = True) -> int | None:
    nh = [norm(h) for h in headers]
    na = [norm(a) for a in aliases]
    for alias in na:
        for i, current in enumerate(nh):
            if current == alias:
                return i
    for alias in na:
        at = set(alias.split())
        for i, current in enumerate(nh):
            if at and at.issubset(set(current.split())):
                return i
    if required:
        raise KeyError(f"Coluna nao encontrada: {aliases}. Cabecalhos: {list(headers)}")
    return None


def row_value(row: Sequence[Any], index: int | None) -> Any:
    return row[index] if index is not None and index < len(row) else ""


@dataclass
class ClientCatalog:
    label: str = ""
    by_pair: dict[tuple[str, str], list[str]] = field(default_factory=lambda: defaultdict(list))
    by_cpf: dict[str, list[str]] = field(default_factory=lambda: defaultdict(list))
    by_name: dict[str, list[str]] = field(default_factory=lambda: defaultdict(list))
    used_codes: set[int] = field(default_factory=set)


def _append_unique(values: list[str], code: str) -> None:
    if code not in values:
        values.append(code)


def add_client_to_catalog(catalog: ClientCatalog, code: Any, name: Any, cpf: Any) -> None:
    code_text = normalize_id(code)
    if not code_text:
        return
    if code_text.isdigit() and 100000 <= int(code_text) <= 999999:
        catalog.used_codes.add(int(code_text))
    cpf_key = digits(cpf)
    name_key = normalize_name(name)
    if cpf_key and name_key:
        _append_unique(catalog.by_pair[(cpf_key, name_key)], code_text)
    if cpf_key:
        _append_unique(catalog.by_cpf[cpf_key], code_text)
    if name_key:
        _append_unique(catalog.by_name[name_key], code_text)


def read_table(path: Path) -> tuple[list[Any], list[list[Any]], str, str]:
    if path.suffix.lower() in {".csv", ".txt"}:
        headers, rows, delimiter, encoding = read_csv(path)
        return headers, rows, delimiter, encoding
    if path.suffix.lower() == ".xlsx":
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
            return headers, rows, ";", "utf-8-sig"
        finally:
            wb.close()
    raise ValueError(f"Formato de planilha de clientes nao suportado: {path.name}")


def load_client_catalog(path: Path, label: str = "") -> ClientCatalog:
    headers, rows, _, _ = read_table(path)
    code_i = find_header(headers, ["Codigo do Cliente", "Codigo Cliente", "Codigo (max 6 digitos)"])
    name_i = find_header(headers, ["Nome", "Nome do Cliente", "Cliente"], required=False)
    cpf_i = find_header(headers, ["CPF"], required=False)
    catalog = ClientCatalog(label=label or path.name)
    for row in rows:
        add_client_to_catalog(
            catalog,
            row_value(row, code_i),
            row_value(row, name_i),
            row_value(row, cpf_i),
        )
    return catalog


def match_client(catalog: ClientCatalog, cpf: Any, name: Any) -> tuple[str, str]:
    cpf_key = digits(cpf)
    name_key = normalize_name(name)

    # O tratamento manual preserva o cliente quando CPF + nome identificam a
    # mesma pessoa, mesmo que existam duplicidades historicas de cadastro. A
    # ordem da planilha fonte e usada de forma deterministica.
    pair_codes = catalog.by_pair.get((cpf_key, name_key), []) if cpf_key and name_key else []
    if pair_codes:
        return pair_codes[0], "CPF+NOME"

    cpf_codes = catalog.by_cpf.get(cpf_key, []) if cpf_key else []
    name_codes = catalog.by_name.get(name_key, []) if name_key else []
    if len(cpf_codes) == 1:
        return cpf_codes[0], "CPF"
    if cpf_codes and name_codes:
        name_set = set(name_codes)
        intersection = [code for code in cpf_codes if code in name_set]
        if intersection:
            return intersection[0], "CPF+NOME"
    if len(name_codes) == 1:
        return name_codes[0], "NOME"
    if len(cpf_codes) > 1:
        return "", "AMBIGUO_CPF"
    if len(name_codes) > 1:
        return "", "AMBIGUO_NOME"
    return "", "NAO_LOCALIZADO"


def match_client_sources(catalogs: Sequence[ClientCatalog], cpf: Any, name: Any) -> tuple[str, str]:
    unresolved: list[str] = []
    for catalog in catalogs:
        code, method = match_client(catalog, cpf, name)
        if code:
            return code, f"{catalog.label}:{method}"
        unresolved.append(method)
    if any(method.startswith("AMBIGUO") for method in unresolved):
        return "", "AMBIGUO"
    return "", "NAO_LOCALIZADO"


def next_client_code(used_codes: set[int]) -> str:
    code = 100000
    while code in used_codes:
        code += 1
    if code > 999999:
        raise ValueError("Sem codigo de cliente disponivel entre 100000 e 999999")
    used_codes.add(code)
    return str(code)


def resolve_sales_plan_clients(directory: Path) -> Path | None:
    accepted = {
        "planilhatratadaclientevendaplano",
        "planilhatratadaclientesvendaplano",
        "planilhatratadaclientevendaplanos",
        "planilhatratadaclientesvendaplanos",
    }
    candidates = [
        path for path in directory.iterdir()
        if path.is_file()
        and path.suffix.lower() in {".csv", ".xlsx"}
        and _layout_key(path.stem) in accepted
    ]
    return newest(candidates).resolve() if candidates else None


@dataclass
class PaymentRule:
    code: str
    destination: str = ""


def _select_fp_sheet(wb: Any) -> Any:
    aliases = {"fp", "formas de pagamento", "forma de pagamento", "formas pagamento", "pagamento"}
    normalized = [(norm(name), name) for name in wb.sheetnames]
    for wanted in aliases:
        nw = norm(wanted)
        for current, original in normalized:
            if current == nw or nw in current or current in nw:
                return wb[original]
    raise KeyError(f"Aba FP/Formas de Pagamento nao encontrada. Abas: {wb.sheetnames}")


def load_payment_rules(path: Path) -> dict[str, PaymentRule]:
    if path.suffix.lower() == ".xls":
        raise ValueError(
            "O DE-PARA esta em .xls. Este script usa openpyxl; salve uma copia do mesmo DE-PARA como .xlsx "
            "sem alterar suas abas e execute novamente."
        )
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = _select_fp_sheet(wb)
        headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        source_i = find_header(headers, ["Forma de Pagamento", "Descricao", "Origem", "Forma"], required=False)
        code_i = find_header(headers, ["Codigo Forma de Pagamento", "Codigo", "Cod", "Destino"], required=False)
        account_i = find_header(headers, ["Codigo da Conta Destino", "Conta Destino", "Conta"], required=False)
        if source_i is None or code_i is None:
            # fallback conservador: primeiras duas colunas preenchidas
            nonempty = [i for i, h in enumerate(headers) if clean_text(h)]
            if len(nonempty) < 2:
                raise ValueError("Aba FP do DE-PARA nao possui colunas suficientes")
            source_i, code_i = nonempty[:2]
        rules: dict[str, PaymentRule] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            source = norm(row_value(row, source_i))
            code = clean_text(row_value(row, code_i))
            if not source or not code:
                continue
            destination = clean_text(row_value(row, account_i)) if account_i is not None else ""
            rules[source] = PaymentRule(code=code, destination=destination)
        if not rules:
            raise ValueError("Nenhuma forma de pagamento valida encontrada na aba FP do DE-PARA")
        return rules
    finally:
        wb.close()


def payment_lookup(rules: dict[str, PaymentRule], value: Any) -> PaymentRule | None:
    key = norm(value)
    if not key:
        return None
    if key in rules:
        return rules[key]
    # equivalencias seguras, ainda usando o DE-PARA como fonte do codigo
    replacements = {
        "credito": ["cartao credito", "credito"],
        "debito": ["cartao debito", "debito"],
        "dinheiro": ["dinheiro"],
        "pix": ["pix"],
    }
    for family, candidates in replacements.items():
        if family in key:
            for candidate in candidates:
                if candidate in rules:
                    return rules[candidate]
            for rule_key, rule in rules.items():
                if family in rule_key:
                    return rule
    return None


# ---------------------------------------------------------------------------
# Regras da importacao
# ---------------------------------------------------------------------------
SOURCE_ALIASES = {
    "account": ["Conta"],
    "sale": ["Venda"],
    "billing": ["Dt. Faturamento Venda", "Data Faturamento Venda"],
    "employee": ["Colaborador"],
    "sale_status": ["Status Venda"],
    "client": ["Cliente"],
    "cpf": ["CPF"],
    "due": ["Data de vencimento", "Vencimento"],
    "paid": ["Data de pagamento", "Pagamento"],
    "payment_expected": ["Forma de pagamento Prevista"],
    "payment_done": ["Forma de pagamento Realizada"],
    "amount": ["Valor total", "Valor"],
    "payment_status": ["Status do pagamento"],
    "delinquent": ["Inadimplente"],
}

MODEL_HEADERS_EXPECTED = [
    "Codigo (max 6 digitos)",
    "Data Lancamento(DD/MM/AAAA Obrigatorio)",
    "Data Competencia(DD/MM/AAAA Obrigatorio)",
    "Data Vencimento(DD/MM/AAAA Obrigatorio se Confirmado)",
    "Data Confirmacao(DD/MM/AAAA Obrigatorio se Confirmado)",
    "Codigo do Cliente",
    "Codigo do Cliente Pagador",
    "Codigo da Conta Origem(Obrigatorio)",
    "Codigo da Conta Destino(Obrigatorio)",
    "Codigo Forma de Pagamento(Obrigatorio)",
    "Aliquota",
    "Valor Titulo(Obrigatorio)",
    "Valor Juros/Multa",
    "Valor Recebido(Obrigatorio se Confirmado)",
    "Confirmado(S/N)",
    "Observacao",
    "Num Doc",
    "Num Cheque",
    "Num NF",
]


@dataclass
class Issue:
    level: str
    line: int
    account: str
    category: str
    detail: str


def eligible(values: dict[str, Any]) -> tuple[bool, str]:
    expected = norm(values["payment_expected"])
    payment_status = norm(values["payment_status"])
    sale_status = norm(values["sale_status"])
    if "pagolivre" in expected or "recorrente" in expected:
        return False, "RECORRENCIA_PAGOLIVRE"
    if payment_status in {"cancelado", "agendado"}:
        return False, f"PAGAMENTO_{payment_status.upper()}"
    if payment_status == "paga":
        return True, ""
    if payment_status == "nao paga" and sale_status == "pendente a pagamento":
        return True, ""
    return False, "STATUS_NAO_IMPORTAVEL"


def default_destination(payment_code: str) -> str:
    key = norm(payment_code)
    # tratamento observado no layout de referencia; o DE-PARA tem prioridade
    if "credito" in key or key.startswith("cc"):
        return "1.1.1.002"
    return "1.1.1.003"


def build_observation(v: dict[str, Any], import_date: str) -> str:
    parts = [
        f"Conta: {clean_text(v['account'])}",
        f"Venda: {clean_text(v['sale'])}",
        f"Data Faturamento: {date_out(v['billing'])}",
        f"Colaborador: {clean_text(v['employee'])}",
        f"Status Venda: {clean_text(v['sale_status'])}",
        f"Forma de pagamento prevista: {clean_text(v['payment_expected'])}",
        f"Forma de pagamento Realizada: {clean_text(v['payment_done'])}",
        f"Status do pagamento: {clean_text(v['payment_status'])}",
        f"Inadimplente: {clean_text(v['delinquent'])}",
    ]
    # nome/CPF, vencimento, pagamento e valor ja possuem destino na importacao
    parts = [p for p in parts if p.split(":", 1)[1].strip()]
    parts.append(f"Importação {import_date}")
    return " | ".join(parts)


def required_model_indexes(headers: Sequence[str]) -> list[int]:
    return [
        i for i, h in enumerate(headers)
        if "obrigatorio" in norm(h) and "obrigatorio se" not in norm(h)
    ]


def write_csv_atomic(
    path: Path,
    headers: Sequence[Any],
    rows: Iterable[Sequence[Any]],
    *,
    delimiter: str = ";",
    encoding: str = "utf-8-sig",
) -> None:
    if not path.parent.is_dir():
        raise FileNotFoundError(f"Pasta de destino nao existe: {path.parent}")
    temp = path.with_name(path.name + ".tmp")
    try:
        with temp.open("w", encoding=encoding, newline="") as f:
            writer = csv.writer(f, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
            writer.writerow([clean_text(h) for h in headers])
            for row in rows:
                writer.writerow([clean_text(v) for v in row])
        temp.replace(path)
    finally:
        if temp.exists():
            temp.unlink()


def build_client_row(headers: Sequence[Any], code: str, name: str, cpf: str, import_date: str) -> list[str]:
    normalized = [norm(h) for h in headers]

    def first(predicate: Any) -> int | None:
        return next((i for i, header in enumerate(normalized) if predicate(header, i)), None)

    code_i = first(lambda h, i: "codigo" in h.split() and ("cliente" in h.split() or "max" in h.split() or i == 0))
    name_i = first(lambda h, i: "nome" in h.split() and "mae" not in h.split() and "pai" not in h.split())
    cpf_i = first(lambda h, i: "cpf" in h.split())
    obs_i = first(lambda h, i: "observacao" in h.split())
    status_i = first(lambda h, i: h.startswith("status"))
    origin_i = first(lambda h, i: "tipo" in h.split() and "origem" in h.split())
    if code_i is None or name_i is None:
        raise ValueError("A planilha/modelo de clientes precisa conter as colunas Codigo e Nome")

    row = ["" for _ in headers]
    row[code_i] = code
    row[name_i] = name
    if cpf_i is not None:
        row[cpf_i] = cpf
    if obs_i is not None:
        row[obs_i] = f"Cliente complementar gerado por Contas a Receber | Importação {import_date}"
    if status_i is not None:
        row[status_i] = "Leads"
    if origin_i is not None:
        row[origin_i] = "Parcerias"
    return row


def append_sales_plan_clients(
    target: Path,
    model_path: Path | None,
    generated: list[tuple[str, str, str]],
) -> int:
    """Inclui novos clientes e normaliza todo o complementar pelas regras de Clientes."""
    import_date = date.today().strftime("%d/%m/%Y")
    if target.exists():
        headers, existing_rows, delimiter, encoding = read_table(target)
    else:
        if not generated:
            return 0
        if model_path is None or not model_path.is_file():
            raise FileNotFoundError(
                "modeloImportacaoCliente.csv e obrigatorio para criar "
                "planilhaTratadaClienteVendaPlano.csv"
            )
        headers, _, delimiter, encoding = read_table(model_path)
        existing_rows = []
        target = target.with_suffix(".csv")

    normalized = [norm(h) for h in headers]
    def first(predicate: Any) -> int | None:
        return next((i for i, header in enumerate(normalized) if predicate(header, i)), None)
    status_i = first(lambda h, i: h.startswith("status"))
    origin_i = first(lambda h, i: "tipo" in h.split() and "origem" in h.split())
    obs_i = first(lambda h, i: "observacao" in h.split())
    mobile_i = first(lambda h, i: "celular" in h.split() and "ddi" not in h.split() and "2" not in h.split())
    cpf_i = first(lambda h, i: "cpf" in h.split())
    ddi_mobile_i = first(lambda h, i: "ddi" in h.split() and "celular" in h.split() and "2" not in h.split())

    def ensure_width(row: Sequence[Any]) -> list[str]:
        result = [clean_text(value) for value in list(row)[:len(headers)]]
        if len(result) < len(headers):
            result.extend([""] * (len(headers) - len(result)))
        return result

    def append_import(text: Any) -> str:
        cleaned = clean_text(text)
        cleaned = re.sub(
            r"(?i)(?:^|\s*\|\s*)importa[cç][aã]o\s+\d{2}/\d{2}/\d{4}(?=\s*\||$)",
            " ", cleaned,
        )
        cleaned = clean_text(cleaned.strip(" |"))
        return " | ".join(part for part in (cleaned, f"Importação {import_date}") if part)

    normalized_existing: list[list[str]] = []
    for raw_row in existing_rows:
        row = ensure_width(raw_row)
        if status_i is not None and not row[status_i]:
            row[status_i] = "Leads"
        if origin_i is not None and not row[origin_i]:
            row[origin_i] = "Parcerias"
        if obs_i is not None:
            row[obs_i] = append_import(row[obs_i])
        if mobile_i is not None and row[mobile_i]:
            d = digits(row[mobile_i])
            if d.startswith("55") and len(d) in {12,13}:
                d = d[2:]
            if len(d) == 10:
                d = d[:2] + "9" + d[2:]
            row[mobile_i] = f"({d[:2]}){d[2:7]}-{d[7:]}" if len(d) == 11 else ""
            if ddi_mobile_i is not None and row[mobile_i]:
                row[ddi_mobile_i] = "55"
        if cpf_i is not None and row[cpf_i]:
            d = digits(row[cpf_i])
            if len(d) in {9,10}:
                d = d.zfill(11)
            row[cpf_i] = f"{d[:3]}.{d[3:6]}.{d[6:9]}-{d[9:]}" if len(d) == 11 else ""
        normalized_existing.append(row)

    new_rows = [build_client_row(headers, code, name, cpf, import_date) for code, name, cpf in generated]
    required = required_model_indexes([clean_text(h) for h in headers])
    for row, item in zip(new_rows, generated):
        code, name, _ = item
        missing = [clean_text(headers[i]) for i in required if not clean_text(row[i])]
        if not clean_text(code) or not clean_text(name) or missing:
            detail = ", ".join(missing) if missing else "Codigo/Nome"
            raise ValueError(f"Cliente complementar incompleto: codigo {code}; faltando {detail}")

    if target.suffix.lower() == ".xlsx":
        wb = load_workbook(target)
        ws = wb[wb.sheetnames[0]]
        # Reescreve linhas existentes normalizadas e depois adiciona as novas.
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for row in [*normalized_existing, *new_rows]:
            ws.append(row)
        temp = target.with_name(target.stem + ".tmp.xlsx")
        wb.save(temp)
        wb.close()
        temp.replace(target)
    else:
        write_csv_atomic(
            target,
            headers,
            [*normalized_existing, *new_rows],
            delimiter=delimiter,
            encoding=encoding,
        )
    return len(new_rows)


def write_validation(path: Path, issues: list[Issue]) -> None:
    if not issues:
        if path.exists():
            path.unlink()
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Validacao"
    ws.append(["Nivel", "Linha origem", "Conta", "Tipo", "Detalhe"])
    for issue in issues:
        ws.append([issue.level, issue.line, issue.account, issue.category, issue.detail])
    ws.freeze_panes = "A2"
    for col, width in {"A": 12, "B": 14, "C": 14, "D": 34, "E": 90}.items():
        ws.column_dimensions[col].width = width
    temp = path.with_name(path.name + ".tmp.xlsx")
    wb.save(temp)
    wb.close()
    temp.replace(path)


# ---------------------------------------------------------------------------
# Processamento principal
# ---------------------------------------------------------------------------
def process(
    extraction: Path,
    model: Path,
    clients_path: Path,
    sales_plan_clients_path: Path,
    depara: Path,
    output: Path,
    client_model: Path | None,
    validation: Path,
) -> tuple[int, int, int, int, dict[str, int], int]:
    model_headers, _, _, _ = read_csv(model)
    if len(model_headers) != 19:
        raise ValueError(f"Modelo de Contas a Receber deve ter 19 colunas; encontrado {len(model_headers)}")
    for expected in ("Codigo", "Data Lancamento", "Data Competencia", "Data Vencimento", "Codigo do Cliente", "Valor Titulo", "Confirmado"):
        if find_header(model_headers, [expected], required=False) is None:
            raise ValueError(f"Modelo invalido: coluna relacionada a '{expected}' nao encontrada")

    sheet, headers, _, header_row = read_xlsx_headers(extraction)
    indexes = {key: find_header(headers, aliases) for key, aliases in SOURCE_ALIASES.items()}
    source_rows: list[tuple[int, dict[str, Any]]] = []
    skipped = 0
    skipped_reasons: dict[str, int] = defaultdict(int)
    source_total = 0
    with XlsxReader(extraction) as xlsx:
        for line, raw in xlsx.iter_rows(sheet, min_row=header_row + 1, max_col=max(120, len(headers))):
            source_total += 1
            values = {key: row_value(raw, idx) for key, idx in indexes.items()}
            ok, reason = eligible(values)
            if ok:
                source_rows.append((line, values))
            else:
                skipped += 1
                skipped_reasons[reason] += 1

    primary_clients = load_client_catalog(clients_path, clients_path.name)
    secondary_exists = sales_plan_clients_path.is_file()
    secondary_clients = (
        load_client_catalog(sales_plan_clients_path, sales_plan_clients_path.name)
        if secondary_exists
        else ClientCatalog(label=sales_plan_clients_path.name)
    )
    client_sources = [primary_clients, secondary_clients]
    used_codes = set(primary_clients.used_codes) | set(secondary_clients.used_codes)
    pay_rules = load_payment_rules(depara)
    import_date = date.today().strftime("%d/%m/%Y")

    prepared: list[tuple[int, dict[str, Any], str, str, str]] = []
    generated_by_identity: dict[tuple[str, str], tuple[str, str, str]] = {}
    issues: list[Issue] = []

    for line, values in source_rows:
        account = normalize_id(values["account"])
        sale_code = normalize_id(values["sale"])
        if not sale_code:
            issues.append(Issue("ERRO", line, account, "VENDA_SEM_CODIGO", "A coluna Venda esta vazia."))
            continue

        client_code, _ = match_client_sources(client_sources, values["cpf"], values["client"])
        if not client_code:
            cpf_key = digits(values["cpf"])
            name_key = normalize_name(values["client"])
            if not name_key:
                issues.append(Issue(
                    "ERRO",
                    line,
                    account,
                    "CLIENTE_SEM_DADOS_SUFICIENTES",
                    "Nome ausente; nao foi possivel localizar ou criar o cliente.",
                ))
                continue
            identity = (cpf_key, name_key)
            if identity not in generated_by_identity:
                code = next_client_code(used_codes)
                generated_by_identity[identity] = (
                    code,
                    clean_text(values["client"]),
                    clean_text(values["cpf"]),
                )
                add_client_to_catalog(secondary_clients, code, values["client"], values["cpf"])
            client_code = generated_by_identity[identity][0]

        payment_name = values["payment_done"] if clean_text(values["payment_done"]) else values["payment_expected"]
        rule = payment_lookup(pay_rules, payment_name)
        if rule is None:
            issues.append(Issue("ERRO", line, account, "FORMA_PAGAMENTO_SEM_DEPARA", clean_text(payment_name)))
            continue
        prepared.append((line, values, account, client_code, rule.code))

    generated = list(generated_by_identity.values())
    added_clients = append_sales_plan_clients(sales_plan_clients_path, client_model, generated)

    rows_out: list[list[str]] = []
    required = required_model_indexes(model_headers)

    for line, values, account, client_code, payment_code in prepared:
        status = norm(values["payment_status"])
        confirmed = "S" if status == "paga" else "N"
        paid_date = date_out(values["paid"])
        billing_date = date_out(values["billing"])
        due_date = date_out(values["due"])
        launch = paid_date or billing_date or due_date
        competence = launch
        confirmation = paid_date if confirmed == "S" else ""
        amount = money(values["amount"])
        received = amount if confirmed == "S" else ""
        rule = payment_lookup(
            pay_rules,
            values["payment_done"] if clean_text(values["payment_done"]) else values["payment_expected"],
        )
        assert rule is not None
        destination = clean_text(rule.destination) or default_destination(payment_code)
        # Codigo de importacao deve ser unico por titulo, mesmo quando uma venda
        # possui varias parcelas/contas. O codigo original da venda permanece na
        # Observacao para rastreabilidade.
        import_code = str(100000 + len(rows_out))
        row = [
            import_code,
            launch,
            competence,
            due_date,
            confirmation,
            client_code,
            client_code,
            "3.1.1.002",
            destination,
            payment_code,
            "",
            amount,
            "",
            received,
            confirmed,
            build_observation(values, import_date),
            "",
            "",
            "",
        ]
        missing = [model_headers[i] for i in required if not clean_text(row[i])]
        if confirmed == "S" and (not due_date or not confirmation or not received):
            missing.extend([x for x in ("Data Vencimento", "Data Confirmacao", "Valor Recebido") if x not in missing])
        if missing:
            issues.append(Issue("ERRO", line, account, "CAMPOS_OBRIGATORIOS", ", ".join(missing)))
            continue
        rows_out.append(row)

    codes = [clean_text(row[0]) for row in rows_out]
    if len(codes) != len(set(codes)):
        raise ValueError("Codigo da coluna A de Contas a Receber nao ficou unico.")
    if any(not re.fullmatch(r"\d{1,6}", code) for code in codes):
        raise ValueError("Codigo da coluna A de Contas a Receber deve ser numerico e ter no maximo 6 digitos.")

    write_csv_atomic(output, model_headers, rows_out)
    write_validation(validation, issues)
    return len(rows_out), skipped, len(issues), added_clients, dict(skipped_reasons), source_total


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Tratamento de Contas a Receber - Laser Rosa")
    p.add_argument("--entrada", type=Path, help="Extracao .xlsx; se omitida, identificada pelo layout em entrada")
    p.add_argument("--modelo", type=Path, help="modeloImportacaoContasReceber.csv")
    p.add_argument("--clientes", type=Path, help="planilhaTratadaCliente.csv; padrao em saida")
    p.add_argument(
        "--clientes-venda-plano",
        type=Path,
        help="planilhaTratadaClienteVendaPlano.csv; segunda fonte e destino de novos clientes",
    )
    p.add_argument("--de-para", dest="depara", type=Path, help="Arquivo unico DE-PARA.xlsx")
    p.add_argument("--modelo-clientes", type=Path, help="modeloImportacaoCliente.csv, opcional se todos clientes existirem")
    p.add_argument("--saida", type=Path, help="CSV final")
    p.add_argument("--validacao", type=Path, help="XLSX de pendencias")
    return p


def resolve_arg(path: Path | None, directory: Path) -> Path | None:
    if path is None:
        return None
    if path.is_absolute():
        return path.resolve()
    if path.parent == Path("."):
        return (directory / path).resolve()
    return (PROJECT_ROOT / path).resolve()


def main() -> int:
    args = build_parser().parse_args()
    try:
        configure_project_layout()
        extraction = discover_extraction(INPUT_DIR, resolve_arg(args.entrada, INPUT_DIR))
        model = resolve_arg(args.modelo, INPUT_DIR) or resolve_variant("modeloImportacaoContasReceber", {".csv"}, INPUT_DIR)
        clients_path = resolve_arg(args.clientes, OUTPUT_DIR) or resolve_variant("planilhaTratadaCliente", {".csv", ".xlsx"}, OUTPUT_DIR)
        explicit_sales_clients = resolve_arg(args.clientes_venda_plano, OUTPUT_DIR)
        sales_plan_clients = (
            explicit_sales_clients
            or resolve_sales_plan_clients(OUTPUT_DIR)
            or (OUTPUT_DIR / "planilhaTratadaClienteVendaPlano.csv")
        )
        depara = resolve_arg(args.depara, INPUT_DIR) or resolve_depara(INPUT_DIR)
        client_model = resolve_arg(args.modelo_clientes, INPUT_DIR) or resolve_variant("modeloImportacaoCliente", {".csv", ".xlsx"}, INPUT_DIR, required=False)
        output = resolve_arg(args.saida, OUTPUT_DIR) or (OUTPUT_DIR / "planilhaTratadaContasReceber.csv")
        validation = resolve_arg(args.validacao, OUTPUT_DIR) or (OUTPUT_DIR / "validacaoContasReceber.xlsx")
        assert model and clients_path and sales_plan_clients and depara and output and validation

        exported, skipped, findings, added_clients, skipped_reasons, source_total = process(
            extraction,
            model,
            clients_path,
            sales_plan_clients,
            depara,
            output,
            client_model,
            validation,
        )

        print(f"OK [v{VERSION}]: {exported} conta(s) exportada(s) de {source_total} linha(s) da extracao.")
        if skipped:
            labels = {
                "RECORRENCIA_PAGOLIVRE": "Recorrencia/PagoLivre",
                "PAGAMENTO_CANCELADO": "Pagamento cancelado",
                "PAGAMENTO_AGENDADO": "Pagamento agendado",
                "STATUS_NAO_IMPORTAVEL": "Status nao importavel",
            }
            parts = [
                f"{labels.get(reason, reason)}: {count}"
                for reason, count in sorted(skipped_reasons.items(), key=lambda item: (-item[1], item[0]))
            ]
            print(
                f"Fora desta importacao por regra esperada: {skipped} linha(s)"
                + (" | " + " | ".join(parts) if parts else "")
            )
        if added_clients:
            print(
                f"ATENCAO: {added_clients} novo(s) cliente(s) incluido(s) em "
                f"{sales_plan_clients.name}; importe essa planilha antes de Contas a Receber."
            )
        if findings:
            print(f"ATENCAO: {findings} pendencia(s) em {validation.name}.")
        return 0
    except Exception as exc:
        print(f"ERRO: {type(exc).__name__}: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
